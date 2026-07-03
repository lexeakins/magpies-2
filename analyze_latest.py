"""
Analyze the latest Magpie run from local outputs.

This intentionally reviews source identity vs. output evidence. It does not
trust provider confidence as ground truth.
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DB_PATH = ROOT / "data" / "magpie.db"
OUTPUT_DIR = ROOT / "data" / "outputs"


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Results"] if "Results" in wb.sheetnames else wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(headers, list(r))) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return rows


def latest_jobs(limit: int = 10) -> list[dict]:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    rows = [
        dict(r)
        for r in con.execute(
            """
            SELECT id, experiment_name, status, created_at, output_file,
                   haiku_calls, perplexity_calls, haiku_cost_usd,
                   perplexity_cost_usd, cost_usd
            FROM jobs
            ORDER BY created_at DESC
            LIMIT ?
            """,
            (limit,),
        )
    ]
    con.close()
    return rows


def find_output(job: dict) -> Path | None:
    db_value = job.get("output_file")
    db_path = Path(db_value) if db_value else None
    if db_path and db_path.exists() and db_path.is_file():
        return db_path
    matches = sorted(OUTPUT_DIR.glob(f"results_{job['id']}_*.xlsx"), key=lambda p: p.stat().st_mtime)
    return matches[-1] if matches else None


def as_int(value, default=0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def print_run(job: dict, rows: list[dict], path: Path):
    print("=" * 80)
    print(f"JOB {job['id']} | {job.get('experiment_name') or ''}")
    print("=" * 80)
    print(f"Output: {path}")
    print(f"Rows: {len(rows)}")
    print(f"Statuses:        {dict(Counter(r.get('status') for r in rows))}")
    print(f"Confidence tiers:{dict(Counter(r.get('confidence_tier') for r in rows))}")
    print(f"Identity verdict:{dict(Counter(r.get('identity_verdict') for r in rows))}")
    print(f"URL source dist: {dict(Counter(r.get('url_source') for r in rows))}")
    print(f"Stages run dist: {dict(Counter(r.get('stages_run') for r in rows))}")

    has_final = sum(1 for r in rows if r.get("final_url"))
    has_gmaps = sum(1 for r in rows if r.get("gmaps_website"))
    has_perp = sum(1 for r in rows if r.get("perplexity_url"))
    rejected_candidates = [
        r for r in rows
        if not r.get("final_url") and (r.get("gmaps_website") or r.get("perplexity_url"))
    ]
    print()
    print(f"Has gmaps_website:       {has_gmaps}/{len(rows)}")
    print(f"Has perplexity_url:      {has_perp}/{len(rows)}")
    print(f"Accepted final_url:      {has_final}/{len(rows)}")
    print(f"Rejected URL candidates: {len(rejected_candidates)}/{len(rows)}")

    g_scores = [as_int(r.get("gmaps_identity_score")) for r in rows if r.get("gmaps_identity_score") is not None]
    p_scores = [as_int(r.get("perplexity_identity_score")) for r in rows if r.get("perplexity_identity_score") is not None]
    f_scores = [as_int(r.get("final_confidence_score")) for r in rows if r.get("final_confidence_score") is not None]
    print()
    print(f"GMaps identity scores:      {len(g_scores)}/{len(rows)} avg={round(sum(g_scores)/len(g_scores),1) if g_scores else 'N/A'}")
    print(f"Perplexity identity scores: {len(p_scores)}/{len(rows)} avg={round(sum(p_scores)/len(p_scores),1) if p_scores else 'N/A'}")
    print(f"Final deterministic scores: {len(f_scores)}/{len(rows)} avg={round(sum(f_scores)/len(f_scores),1) if f_scores else 'N/A'}")

    rejected_perplexity_candidates = [
        r for r in rows
        if not r.get("final_url")
        and r.get("perplexity_url")
    ]
    weak_final = [
        r for r in rows
        if r.get("final_url") and as_int(r.get("final_confidence_score")) < 85
    ]

    print_section("Rejected Perplexity Candidates", rejected_perplexity_candidates)
    print_section("Accepted Final URLs Below 85", weak_final)
    print_section("Rejected Candidates", rejected_candidates)


def print_section(title: str, rows: list[dict], limit: int = 12):
    print()
    print(title)
    print("-" * len(title))
    if not rows:
        print("  none")
        return
    for r in rows[:limit]:
        candidate = r.get("perplexity_url") or r.get("gmaps_website") or ""
        print(
            f"  {r.get('company','')[:42]:<42} | "
            f"src={r.get('url_source') or '-':<10} "
            f"final={str(r.get('final_url') or '')[:36]:<36} "
            f"cand={str(candidate)[:36]:<36} "
            f"score={r.get('final_confidence_score')} "
            f"verdict={r.get('identity_verdict')} "
            f"reason={str(r.get('identity_reason') or '')[:90]}"
        )
    if len(rows) > limit:
        print(f"  ... {len(rows) - limit} more")


def compare_runs(left_rows: list[dict], right_rows: list[dict], left_label: str, right_label: str):
    left_by_company = {r.get("company"): r for r in left_rows}
    right_by_company = {r.get("company"): r for r in right_rows}
    companies = sorted(set(left_by_company) | set(right_by_company))

    print()
    print("=" * 80)
    print(f"COMPARISON: {left_label} -> {right_label}")
    print("=" * 80)
    changes = []
    for company in companies:
        a = left_by_company.get(company, {})
        b = right_by_company.get(company, {})
        if (a.get("final_url"), a.get("final_confidence_score"), a.get("status")) != (
            b.get("final_url"), b.get("final_confidence_score"), b.get("status")
        ):
            changes.append((company, a, b))

    print(f"Changed rows: {len(changes)}/{len(companies)}")
    for company, a, b in changes[:20]:
        print(f"  {company[:42]:<42}")
        print(f"    before: {a.get('status')} score={a.get('final_confidence_score')} url={a.get('final_url')}")
        print(f"    after : {b.get('status')} score={b.get('final_confidence_score')} url={b.get('final_url')}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--job-id", help="Analyze a specific job id instead of latest completed job.")
    parser.add_argument("--compare-job", help="Compare analyzed job against another job id.")
    args = parser.parse_args()

    jobs = latest_jobs(25)
    if not jobs:
        raise SystemExit("No jobs found.")

    job = None
    if args.job_id:
        job = next((j for j in jobs if j["id"] == args.job_id), None)
    else:
        job = next((j for j in jobs if j["status"] == "done"), jobs[0])
    if not job:
        raise SystemExit(f"Job not found: {args.job_id}")

    out = find_output(job)
    if not out:
        raise SystemExit(f"No output workbook found for job {job['id']}.")
    rows = load_rows(out)
    print_run(job, rows, out)

    if args.compare_job:
        other = next((j for j in jobs if j["id"] == args.compare_job), None)
        if not other:
            raise SystemExit(f"Compare job not found in recent jobs: {args.compare_job}")
        other_out = find_output(other)
        if not other_out:
            raise SystemExit(f"No output workbook found for compare job {other['id']}.")
        compare_runs(load_rows(other_out), rows, other["id"], job["id"])


if __name__ == "__main__":
    main()
