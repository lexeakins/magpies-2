from openpyxl import load_workbook
from collections import Counter

path = r"C:\Users\Alex Eakins\Desktop\Accounts_Receivable_Contacts\magpie_app\data\outputs\results_b8894e14_20260629_171858.xlsx"
wb = load_workbook(path, read_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
rows = [dict(zip(headers, list(r))) for r in ws.iter_rows(min_row=2, values_only=True)]
wb.close()

print("COLUMNS:", headers)
print("TOTAL:", len(rows))
print("TIERS:", dict(Counter(r.get("confidence_tier") for r in rows)))
print("STATUSES:", dict(Counter(r.get("status") for r in rows)))
found = sum(1 for r in rows if str(r.get("gmaps_found","")).lower() in ("yes","true","1"))
print("GMAPS_FOUND:", found)
print("ATTEMPTS:", dict(Counter(r.get("scrape_attempts") for r in rows)))
scores = [r.get("confidence_score") for r in rows if r.get("confidence_score") is not None]
if scores:
    print("SCORES min/max/avg:", min(scores), max(scores), round(sum(scores)/len(scores),1))
print("HAS_ADDRESS:", sum(1 for r in rows if r.get("gmaps_address")))
print("HAS_WEBSITE:", sum(1 for r in rows if r.get("gmaps_website")))
print("HAS_LISTING:", sum(1 for r in rows if r.get("gmaps_listing_name")))
print("LOC_MATCH:", dict(Counter(r.get("gmaps_location_match") for r in rows)))

print()
print("=== SAMPLE validation_signals (first 8) ===")
for r in rows[:8]:
    print(" ", repr(r.get("validation_signals")))

print()
print("=== SAMPLE haiku_reasoning (first 8) ===")
for r in rows[:8]:
    print(" ", repr(r.get("haiku_reasoning")))

print()
print("=== SAMPLE gmaps_address (first 8) ===")
for r in rows[:8]:
    print(" ", repr(r.get("gmaps_address")))

print()
print("=== LOW CONFIDENCE (first 8) ===")
for r in [x for x in rows if x.get("confidence_tier") == "Low"][:8]:
    print(" ", r.get("company"), "| score=", r.get("confidence_score"),
          "| status=", r.get("status"), "| sig=", r.get("validation_signals"))

print()
print("=== HIGH CONFIDENCE (first 5) ===")
for r in [x for x in rows if x.get("confidence_tier") == "High"][:5]:
    print(" ", r.get("company"), "| score=", r.get("confidence_score"),
          "| sig=", r.get("validation_signals"))
