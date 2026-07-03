from openpyxl import load_workbook
from collections import Counter, defaultdict

path = r"C:\Users\Alex Eakins\Desktop\Accounts_Receivable_Contacts\magpie_app\data\outputs\results_b2bf6f83_20260629_183024.xlsx"
wb = load_workbook(path, read_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
rows = [dict(zip(headers, list(r))) for r in ws.iter_rows(min_row=2, values_only=True)]
wb.close()

print("TOTAL ROWS:", len(rows))
print("TIERS:", dict(Counter(r.get("confidence_tier") for r in rows)))
print("STATUSES:", dict(Counter(r.get("status") for r in rows)))
print("ATTEMPTS:", dict(Counter(r.get("scrape_attempts") for r in rows)))
print("LOC_MATCH:", dict(Counter(r.get("gmaps_location_match") for r in rows)))

# Rows that hit max retries - what did they get?
max_retry_rows = [r for r in rows if str(r.get("status","")).startswith("max_retries")]
print(f"\nMAX RETRY ROWS: {len(max_retry_rows)}")
print("  gmaps_found on max_retry rows:", Counter(str(r.get("gmaps_found")) for r in max_retry_rows))
print("  has website on max_retry rows:", sum(1 for r in max_retry_rows if r.get("gmaps_website")))

# Rows that had 3 attempts - did their scrape result change attempt to attempt?
# We can't see per-attempt history, but we can see the final state
three_attempt = [r for r in rows if r.get("scrape_attempts") == 3]
print(f"\n3-ATTEMPT ROWS ({len(three_attempt)}):")
print("  final status:", Counter(r.get("status") for r in three_attempt))
print("  has website:", sum(1 for r in three_attempt if r.get("gmaps_website")))
print("  loc_match:", Counter(r.get("gmaps_location_match") for r in three_attempt))

# What's missing that we could get
print("\n=== WHAT WE'RE LEAVING ON THE TABLE ===")
has_website = [r for r in rows if r.get("gmaps_website")]
has_phone   = [r for r in rows if r.get("gmaps_phone")]
has_address = [r for r in rows if r.get("gmaps_address")]
has_listing = [r for r in rows if r.get("gmaps_listing_name")]
no_website  = [r for r in rows if not r.get("gmaps_website")]
print(f"Has website:       {len(has_website)}/{len(rows)}")
print(f"Has phone:         {len(has_phone)}/{len(rows)}")
print(f"Has address:       {len(has_address)}/{len(rows)}")
print(f"Has listing name:  {len(has_listing)}/{len(rows)}")
print(f"No website at all: {len(no_website)}")

# Of no-website rows: do they have a phone at least?
no_web_has_phone = [r for r in no_website if r.get("gmaps_phone")]
no_web_has_listing = [r for r in no_website if r.get("gmaps_listing_name")]
print(f"  No-website rows that have a phone:   {len(no_web_has_phone)}")
print(f"  No-website rows that have a listing: {len(no_web_has_listing)}")

# Haiku contribution - what did it actually do?
haiku_rows = [r for r in rows if r.get("sig_site_name") is not None]
print(f"\nHAIKU WAS CALLED ON: {len(haiku_rows)} rows")
if haiku_rows:
    print("  sig_site_name dist:    ", Counter(r.get("sig_site_name") for r in haiku_rows))
    print("  sig_site_location dist:", Counter(r.get("sig_site_location") for r in haiku_rows))
    print("  sig_isn_mention dist:  ", Counter(r.get("sig_isn_mention") for r in haiku_rows))
    print("  sig_disqualifier dist: ", Counter(r.get("sig_disqualifier") for r in haiku_rows))

# Did Haiku change the outcome vs. what stage1 alone would have given?
print("\n=== HAIKU IMPACT ANALYSIS ===")
for r in haiku_rows[:10]:
    sim = r.get("name_similarity", 0) or 0
    loc = r.get("gmaps_location_match", "")
    s1 = 0
    if sim >= 90: s1 += 20
    elif sim >= 70: s1 += 12
    elif sim >= 50: s1 += 6
    if loc == "exact": s1 += 20
    elif loc == "partial": s1 += 10
    elif loc == "unknown": s1 += 5
    print(f"  {str(r.get('company',''))[:35]:<35} | s1={s1:2d} | final={r.get('confidence_score')} | "
          f"site_name={r.get('sig_site_name')} | site_loc={r.get('sig_site_location')} | status={r.get('status')}")

# Medium confidence - what's dragging them down?
medium = [r for r in rows if r.get("confidence_tier") == "Medium"]
print(f"\nMEDIUM CONFIDENCE ({len(medium)}):")
print("  sig_site_name:", Counter(r.get("sig_site_name") for r in medium))
print("  sig_site_location:", Counter(r.get("sig_site_location") for r in medium))
print("  Sample haiku_reasoning:")
for r in medium[:5]:
    print(f"    [{r.get('company','')[:30]}]:", r.get("haiku_reasoning","")[:100])

# Low confidence rows
low = [r for r in rows if r.get("confidence_tier") == "Low"]
print(f"\nLOW CONFIDENCE ({len(low)}):")
for r in low:
    print(f"  {str(r.get('company',''))[:35]:<35} | score={r.get('confidence_score')} | "
          f"site_name={r.get('sig_site_name')} | disq={r.get('sig_disqualifier')} | "
          f"reasoning={str(r.get('haiku_reasoning',''))[:80]}")
