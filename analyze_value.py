from openpyxl import load_workbook
from collections import Counter, defaultdict
import sqlite3, json

T1_PATH = r"C:\Users\Alex Eakins\Desktop\Accounts_Receivable_Contacts\magpie_app\data\outputs\results_4ab3cc86_20260629_201635.xlsx"
T2_PATH = r"C:\Users\Alex Eakins\Desktop\Accounts_Receivable_Contacts\magpie_app\data\outputs\results_211518b6_20260629_202156.xlsx"

def load(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [dict(zip(headers, list(r))) for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return rows

t1 = load(T1_PATH)
t2 = load(T2_PATH)

# Match rows by company for cross-run comparison
t1_by_company = {r['company']: r for r in t1}
t2_by_company = {r['company']: r for r in t2}

print("=" * 65)
print("WHAT DID THE $0.01944 HAIKU SPEND ACTUALLY GET US?")
print("=" * 65)

haiku_rows = [r for r in t2 if r.get('haiku_initial_confidence') is not None]
print(f"\nHaiku ran on {len(haiku_rows)}/30 rows (only when Maps found a website URL)")
print(f"Cost: $0.01944 / {len(haiku_rows)} calls = ${0.01944/len(haiku_rows):.5f} per call")

# Token efficiency
conn = sqlite3.connect(r"C:\Users\Alex Eakins\Desktop\Accounts_Receivable_Contacts\magpie_app\data\magpie.db")
cur = conn.cursor()
cur.execute("SELECT input_tokens, output_tokens, cost_usd, raw_response FROM api_calls WHERE job_id='211518b6' AND provider='anthropic'")
haiku_calls_db = cur.fetchall()
total_in = sum(r[0] for r in haiku_calls_db)
total_out = sum(r[1] for r in haiku_calls_db)
print(f"Tokens: {total_in} input + {total_out} output = {total_in+total_out} total")
print(f"Avg tokens per call: {total_in//len(haiku_calls_db)} in / {total_out//len(haiku_calls_db)} out")

print("\n--- WHAT HAIKU CHANGED vs GMaps-only ---")
improved, degraded, same_tier = 0, 0, 0
tier_order = {'High': 3, 'Medium': 2, 'Low': 1, None: 0}

for r2 in haiku_rows:
    company = r2['company']
    r1 = t1_by_company.get(company)
    if not r1:
        continue
    g_score = r2.get('gmaps_confidence_score', 0) or 0
    h_score = r2.get('haiku_initial_confidence', 0) or 0
    g_tier = r1.get('confidence_tier')
    h_tier = r2.get('confidence_tier')
    if tier_order.get(h_tier,0) > tier_order.get(g_tier,0):
        improved += 1
    elif tier_order.get(h_tier,0) < tier_order.get(g_tier,0):
        degraded += 1
    else:
        same_tier += 1

print(f"  Tier improved (Haiku upgraded confidence):  {improved}")
print(f"  Tier degraded (Haiku caught bad result):    {degraded}")
print(f"  Same tier:                                  {same_tier}")

print("\n--- HAIKU SIGNAL BREAKDOWN (what it actually found) ---")
print(f"  name_match YES:     {sum(1 for r in haiku_rows if r.get('sig_site_name')=='YES')}")
print(f"  name_match PARTIAL: {sum(1 for r in haiku_rows if r.get('sig_site_name')=='PARTIAL')}")
print(f"  name_match NO:      {sum(1 for r in haiku_rows if r.get('sig_site_name')=='NO')}")
print(f"  location YES:       {sum(1 for r in haiku_rows if r.get('sig_site_location')=='YES')}")
print(f"  location NO:        {sum(1 for r in haiku_rows if r.get('sig_site_location')=='NO')}")
print(f"  ISN mention YES:    {sum(1 for r in haiku_rows if r.get('sig_isn_mention')=='YES')}")
print(f"  DISQUALIFIER YES:   {sum(1 for r in haiku_rows if r.get('sig_disqualifier')=='YES')}")

print("\n--- SCORE DELTA: GMaps score vs Haiku-adjusted score ---")
deltas = []
for r in haiku_rows:
    g = r.get('gmaps_confidence_score') or 0
    h = r.get('haiku_initial_confidence') or 0
    deltas.append((h - g, r.get('company','')[:35], g, h, r.get('sig_site_name'), r.get('sig_site_location'), r.get('status')))

deltas.sort(key=lambda x: x[0])
print("  Biggest drops (Haiku said worse than Maps suggested):")
for d, co, g, h, nm, loc, st in deltas[:5]:
    print(f"    {d:+3d} | {co:<35} | gmaps={g} haiku={h} | name={nm} loc={loc} | {st}")
print("  Biggest gains (Haiku confirmed strongly):")
for d, co, g, h, nm, loc, st in deltas[-5:]:
    print(f"    {d:+3d} | {co:<35} | gmaps={g} haiku={h} | name={nm} loc={loc} | {st}")

print("\n--- HAIKU REASONING SAMPLES ---")
for r in haiku_rows[:8]:
    print(f"  [{r.get('status')}] {r.get('company','')[:30]:<30} | score={r.get('haiku_initial_confidence')} | {r.get('haiku_reasoning','')[:80]}")

print("\n--- STOP REASONS (quality check) ---")
stop_reasons = Counter(r[0] for r in haiku_calls_db if r[0])
# Actually stop_reason is in raw_response for us, let me get from job_results
cur.execute("SELECT haiku_initial_stop_reason FROM job_results WHERE job_id='211518b6'")
stops = Counter(r[0] for r in cur.fetchall())
print(f"  {dict(stops)}")

print("\n" + "=" * 65)
print("THE 10 NOT-FOUND RECORDS — WHAT DO WE KNOW?")
print("=" * 65)
not_found = [r for r in t2 if r.get('status') == 'not_found']
for r in not_found:
    g = r.get('gmaps_confidence_score') or 0
    stage = r.get('stages_run','')
    print(f"  {r.get('company','')[:40]:<40} | gmaps_score={g:2d} | stages={stage} | phone={bool(r.get('gmaps_phone'))} | addr={bool(r.get('gmaps_address'))}")

print("\n--- Of these, how many did Perplexity attempt? ---")
perp_attempted_nf = [r for r in not_found if 'perplexity' in (r.get('stages_run') or '')]
print(f"  {len(perp_attempted_nf)}/{len(not_found)} (all failed due to missing API key)")

print("\n" + "=" * 65)
print("CROSS-RUN COMPARISON: SAME 30 ROWS, T1 vs T2")
print("=" * 65)
print(f"\n{'Company':<35} | T1 tier | T2 tier | T1 score | T2 score | Changed?")
print("-" * 85)
for company, r1 in sorted(t1_by_company.items(), key=lambda x: x[0]):
    r2 = t2_by_company.get(company)
    if not r2: continue
    t1_tier = r1.get('confidence_tier','?')
    t2_tier = r2.get('confidence_tier','?')
    t1_sc   = r1.get('final_confidence_score','?')
    t2_sc   = r2.get('final_confidence_score','?')
    changed = "  ** CHANGED **" if t1_tier != t2_tier else ""
    print(f"  {company[:33]:<33} | {str(t1_tier):<7} | {str(t2_tier):<7} | {str(t1_sc):<8} | {str(t2_sc):<8} |{changed}")

print("\n" + "=" * 65)
print("WHERE TO ITERATE NEXT")
print("=" * 65)

# What would Perplexity have run on?
perp_targets = [r for r in t2 if 'perplexity' in (r.get('stages_run') or '')]
print(f"\nPerplexity was triggered on {len(perp_targets)} records:")
no_url_cases = [r for r in perp_targets if not r.get('gmaps_website')]
low_conf_cases = [r for r in perp_targets if r.get('gmaps_website')]
print(f"  {len(no_url_cases)} with NO website from Maps  ← Perplexity's primary job")
print(f"  {len(low_conf_cases)} with a website but Haiku said low confidence")

print(f"\nWith the Perplexity key now fixed, next run on these {len(perp_targets)} records will tell us:")
print(f"  - Can Perplexity find websites for the {len(no_url_cases)} Maps-missing cases?")
print(f"  - Can it correct/confirm the {len(low_conf_cases)} low-confidence URLs?")

print(f"\nHaiku stop reasons (truncation check):")
cur.execute("SELECT haiku_initial_stop_reason, COUNT(*) FROM job_results WHERE job_id='211518b6' GROUP BY haiku_initial_stop_reason")
for row in cur.fetchall():
    print(f"  {row[0]}: {row[1]}")

cur.execute("SELECT AVG(haiku_initial_latency_ms), MIN(haiku_initial_latency_ms), MAX(haiku_initial_latency_ms) FROM job_results WHERE job_id='211518b6' AND haiku_initial_latency_ms IS NOT NULL")
row = cur.fetchone()
if row[0]:
    print(f"\nHaiku latency: avg={int(row[0])}ms  min={row[1]}ms  max={row[2]}ms")

conn.close()
