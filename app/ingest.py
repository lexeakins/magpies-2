import re
from pathlib import Path
from openpyxl import load_workbook
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

# ---------------------------------------------------------------------------
# Known geographic values for column detection and country inference
# ---------------------------------------------------------------------------

US_STATES = {
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho",
    "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana",
    "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada",
    "New Hampshire", "New Jersey", "New Mexico", "New York",
    "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon",
    "Pennsylvania", "Rhode Island", "South Carolina", "South Dakota",
    "Tennessee", "Texas", "Utah", "Vermont", "Virginia", "Washington",
    "West Virginia", "Wisconsin", "Wyoming", "District of Columbia",
}
US_ABBREVS = {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
    "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
    "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
    "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
    "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "DC",
}
CA_PROVINCES = {
    "Ontario", "Quebec", "British Columbia", "Alberta", "Manitoba",
    "Saskatchewan", "Nova Scotia", "New Brunswick",
    "Newfoundland and Labrador", "Prince Edward Island",
    "Northwest Territories", "Yukon", "Nunavut",
}
CA_ABBREVS = {"ON", "QC", "BC", "AB", "MB", "SK", "NS", "NB", "NL", "PE", "NT", "YT", "NU"}

COUNTRY_VALUES = {"united states", "usa", "us", "canada", "ca", "mexico", "mx"}

HEADER_KEYWORDS = ["company", "name", "city", "state", "country", "organization", "business", "account"]
COMPANY_KEYWORDS = ["company", "name", "business", "organization", "account"]
CITY_KEYWORDS = ["city", "town", "municipality", "location"]
STATE_KEYWORDS = ["state", "province", "region"]
COUNTRY_KEYWORDS = ["country", "nation"]


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class Lead:
    company: str
    city: str
    state: str
    country: str
    source_sheet: str
    original_row_idx: int
    source_valid_index: int = 0


@dataclass
class DropRecord:
    sheet: str
    row_idx: int
    raw_data: str
    reason: str


@dataclass
class CleanseReport:
    sheets_found: List[str]
    total_rows_raw: int
    total_rows_valid: int
    dropped_rows: int
    drop_reasons: dict          # reason -> count
    country_distribution: dict  # country -> count
    sample_drops: List[dict]    # first 10 dropped rows
    sample_drops_by_reason: dict = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Country inference helpers
# ---------------------------------------------------------------------------

def _infer_country_from_state(state_val: str) -> Optional[str]:
    if not state_val:
        return None
    sv = state_val.strip()
    if sv in US_STATES or sv.title() in US_STATES or sv.upper() in US_ABBREVS:
        return "United States"
    if sv in CA_PROVINCES or sv.title() in CA_PROVINCES or sv.upper() in CA_ABBREVS:
        return "Canada"
    return None


def _infer_country_from_sheet(sheet_name: str) -> Optional[str]:
    sn = sheet_name.strip().upper()
    if sn in US_ABBREVS or sn in {s.upper() for s in US_STATES}:
        return "United States"
    if sn in {"CAN", "CANADA"} or sn in CA_ABBREVS:
        return "Canada"
    return None


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

def _is_header_row(row: list) -> bool:
    """Return True if the row looks like column labels rather than real data."""
    if not row:
        return False
    text = " ".join(str(v).lower() for v in row if v)
    return sum(1 for kw in HEADER_KEYWORDS if kw in text) >= 2


def _detect_columns(rows: list, has_header: bool) -> dict:
    """
    Returns {"company": idx, "city": idx, "state": idx, "country": idx_or_None}.
    Uses header text if present, otherwise infers from cell content.
    """
    if not rows:
        return {}

    if has_header:
        header = [str(v).lower().strip() if v else "" for v in rows[0]]
        result = {}
        for col_idx, h in enumerate(header):
            if not h:
                continue
            if "company" not in result and any(k in h for k in COMPANY_KEYWORDS):
                result["company"] = col_idx
            elif "city" not in result and any(k in h for k in CITY_KEYWORDS):
                result["city"] = col_idx
            elif "state" not in result and any(k in h for k in STATE_KEYWORDS):
                result["state"] = col_idx
            elif "country" not in result and any(k in h for k in COUNTRY_KEYWORDS):
                result["country"] = col_idx
        return result

    # No header — infer by content
    sample = [r for r in rows[:30] if any(c for c in r if c)]
    if not sample:
        return {}

    num_cols = max(len(r) for r in sample)

    # State column: highest proportion of known state/province values
    state_col, best_state_pct = None, 0.0
    for ci in range(num_cols):
        vals = [str(r[ci]).strip() for r in sample if ci < len(r) and r[ci]]
        if not vals:
            continue
        hits = sum(
            1 for v in vals
            if v in US_STATES or v.upper() in US_ABBREVS
            or v in CA_PROVINCES or v.upper() in CA_ABBREVS
        )
        pct = hits / len(vals)
        if pct > best_state_pct and pct >= 0.4:
            best_state_pct, state_col = pct, ci

    # Country column: values like "United States", "Canada"
    country_col = None
    for ci in range(num_cols):
        if ci == state_col:
            continue
        vals = [str(r[ci]).strip().lower() for r in sample if ci < len(r) and r[ci]]
        if not vals:
            continue
        if sum(1 for v in vals if v in COUNTRY_VALUES) / len(vals) >= 0.5:
            country_col = ci
            break

    # Company column: longest average string length among remaining columns
    used = {state_col, country_col}
    company_col, best_len = None, 0.0
    for ci in range(num_cols):
        if ci in used:
            continue
        vals = [str(r[ci]).strip() for r in sample if ci < len(r) and r[ci]]
        if not vals:
            continue
        avg = sum(len(v) for v in vals) / len(vals)
        if avg > best_len:
            best_len, company_col = avg, ci

    # City column: whatever reasonable column remains
    city_col = None
    used = {state_col, country_col, company_col}
    for ci in range(num_cols):
        if ci not in used:
            vals = [str(r[ci]).strip() for r in sample if ci < len(r) and r[ci]]
            if vals:
                city_col = ci
                break

    result = {}
    if company_col is not None:
        result["company"] = company_col
    if city_col is not None:
        result["city"] = city_col
    if state_col is not None:
        result["state"] = state_col
    if country_col is not None:
        result["country"] = country_col
    return result


# ---------------------------------------------------------------------------
# Main ingestion function
# ---------------------------------------------------------------------------

def ingest_file(file_path: str) -> Tuple[List[Lead], CleanseReport]:
    """
    Load an Excel file (any number of sheets), detect columns, infer country,
    validate rows, and return (valid_leads, cleanse_report).
    """
    wb = load_workbook(file_path, read_only=True)
    sheet_names = wb.sheetnames

    all_leads: List[Lead] = []
    all_drops: List[DropRecord] = []
    total_raw = 0

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]

        # Skip completely empty sheets
        if not rows or all(all(c is None for c in r) for r in rows):
            continue

        has_header = _is_header_row(rows[0])
        data_rows = rows[1:] if has_header else rows
        if not data_rows:
            continue

        col_map = _detect_columns(rows, has_header)

        # If we can't find the required columns, skip the sheet and log drops
        if "company" not in col_map or "city" not in col_map or "state" not in col_map:
            for ri, row in enumerate(data_rows):
                total_raw += 1
                all_drops.append(DropRecord(
                    sheet=sheet_name,
                    row_idx=ri + (2 if has_header else 1),
                    raw_data=_row_preview(row),
                    reason="could_not_detect_columns",
                ))
            continue

        total_raw += len(data_rows)

        for ri, row in enumerate(data_rows):
            company = _cell(row, col_map.get("company"))
            city    = _cell(row, col_map.get("city"))
            state   = _cell(row, col_map.get("state"))
            country = _cell(row, col_map.get("country"))

            # Infer country if missing
            if not country:
                country = _infer_country_from_state(state)
            if not country:
                country = _infer_country_from_sheet(sheet_name)
            if not country:
                country = "United States"

            # Validate required fields
            drop_reason = None
            if not company:
                drop_reason = "missing_company"
            elif not city:
                drop_reason = "missing_city"
            elif not state:
                drop_reason = "missing_state"
            elif len(company) < 2:
                drop_reason = "company_too_short"

            row_num = ri + (2 if has_header else 1)
            if drop_reason:
                all_drops.append(DropRecord(
                    sheet=sheet_name,
                    row_idx=row_num,
                    raw_data=f"{company} | {city} | {state}",
                    reason=drop_reason,
                ))
            else:
                all_leads.append(Lead(
                    company=company.strip(),
                    city=city.strip(),
                    state=state.strip(),
                    country=country.strip(),
                    source_sheet=sheet_name,
                    original_row_idx=row_num,
                    source_valid_index=len(all_leads) + 1,
                ))

    wb.close()

    # Build report
    reason_counts: dict = {}
    for d in all_drops:
        reason_counts[d.reason] = reason_counts.get(d.reason, 0) + 1

    country_dist: dict = {}
    for lead in all_leads:
        country_dist[lead.country] = country_dist.get(lead.country, 0) + 1

    samples_by_reason: dict = {}
    for d in all_drops:
        bucket = samples_by_reason.setdefault(d.reason, [])
        if len(bucket) < 5:
            bucket.append({"sheet": d.sheet, "row": d.row_idx, "data": d.raw_data, "reason": d.reason})

    report = CleanseReport(
        sheets_found=sheet_names,
        total_rows_raw=total_raw,
        total_rows_valid=len(all_leads),
        dropped_rows=len(all_drops),
        drop_reasons=reason_counts,
        country_distribution=country_dist,
        sample_drops=[
            {"sheet": d.sheet, "row": d.row_idx, "data": d.raw_data, "reason": d.reason}
            for d in all_drops[:10]
        ],
        sample_drops_by_reason=samples_by_reason,
    )

    return all_leads, report


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(row: list, idx: Optional[int]) -> Optional[str]:
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


def _row_preview(row: list) -> str:
    return " | ".join(str(v) for v in row if v)[:120]
