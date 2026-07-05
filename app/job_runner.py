"""
Pipeline orchestration.

Stages (each independently toggleable):
  1. GMAPS    — always runs. Produces gmaps_confidence_score.
  2. HAIKU    — validates Maps URL. Fires if haiku_enabled and URL exists.
  3. PERPLEXITY — finds/verifies URL. Fires based on perplexity_trigger.
  4. HAIKU_VAL  — validates Perplexity's URL. Fires if haiku_validation_enabled.
"""

import os, platform, socket, sys, time, uuid, json, queue, re, threading, traceback
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor
from typing import List, Optional
from urllib.parse import urlparse

from .ingest import Lead
from .scraper import scrape_lead
from .validator import fetch_page_snippet, validate_with_haiku
from .perplexity_client import call_perplexity
from .manual_sources import lookup_manual_associations
from .historical_sources import lookup_historical_candidates
from .web_search import lookup_web_candidates
from .bing_maps import lookup_bing_maps_candidates
from .gmaps import setup_driver
from .scorer import (
    name_similarity, gmaps_confidence_score, haiku_confidence_score,
    score_candidate_url, business_identity_overlap
)
from .config import settings
from . import database as db


APP_STARTED_AT = datetime.now(timezone.utc)
APP_MONOTONIC_STARTED = time.monotonic()


# ---------------------------------------------------------------------------
# Status constants
# ---------------------------------------------------------------------------
class Status:
    PENDING   = "pending"
    RUNNING   = "running"
    DONE      = "done"
    CANCELLED = "cancelled"
    ERROR     = "error"


# ---------------------------------------------------------------------------
# Thread-safe in-flight counter
# ---------------------------------------------------------------------------
class _Counter:
    def __init__(self):
        self._v = 0
        self._lock = threading.Lock()
        self._zero = threading.Event()
        self._zero.set()

    def inc(self):
        with self._lock:
            self._v += 1
            self._zero.clear()

    def dec(self):
        with self._lock:
            self._v = max(0, self._v - 1)
            if self._v == 0:
                self._zero.set()

    def wait(self, timeout=7200):
        return self._zero.wait(timeout=timeout)

    def value(self):
        with self._lock:
            return self._v


# ---------------------------------------------------------------------------
# Address parser
# ---------------------------------------------------------------------------
_ADDR_RE = re.compile(r'^(.+?),\s*(.+?),\s*([A-Z]{2})\s*(\S+)?$')

def parse_gmaps_address(addr: str) -> dict:
    if not addr:
        return {"gmaps_street": None, "gmaps_city_scraped": None,
                "gmaps_state_scraped": None, "gmaps_zip": None}
    m = _ADDR_RE.match(addr.strip())
    if m:
        return {"gmaps_street":        m.group(1).strip() or None,
                "gmaps_city_scraped":  m.group(2).strip() or None,
                "gmaps_state_scraped": m.group(3).strip() or None,
                "gmaps_zip":           (m.group(4) or "").strip() or None}
    return {"gmaps_street": addr, "gmaps_city_scraped": None,
            "gmaps_state_scraped": None, "gmaps_zip": None}


def _environment_snapshot() -> dict:
    return {
        "app_started_at": APP_STARTED_AT.isoformat(),
        "app_uptime_sec_at_job_start": round(time.monotonic() - APP_MONOTONIC_STARTED, 1),
        "process_id": os.getpid(),
        "host_name": socket.gethostname(),
        "platform": platform.platform(),
        "python_version": sys.version.split()[0],
        "cwd": os.getcwd(),
    }


def _effective_job_settings(job_settings: dict | None = None) -> dict:
    job_settings = job_settings or {}
    keys = [
        "scrape_workers",
        "validate_workers",
        "job_stall_timeout_seconds",
        "jina_timeout_seconds",
        "max_scrape_errors",
        "haiku_confidence_threshold",
        "gmaps_max_candidates_per_mode",
        "gmaps_strong_stop_score",
        "web_search_max_results",
        "historical_enrichment_min_score",
    ]
    base = settings.pipeline_snapshot()
    return {key: job_settings.get(key, base.get(key)) for key in keys}


# ---------------------------------------------------------------------------
# Job report builder
# ---------------------------------------------------------------------------
def build_report(job: "Job") -> dict:
    results   = job.results
    completed = max(job.completed, 1)
    tiers     = Counter(r.get("confidence_tier") for r in results)
    statuses  = Counter(r.get("status") for r in results)
    identity  = Counter(r.get("identity_verdict") for r in results)
    stage_counts = Counter(e.get("event_type") for e in job.stage_events_snapshot(limit=0))
    active_rows = job.active_rows_snapshot()
    scores    = [r["final_confidence_score"] for r in results
                 if r.get("final_confidence_score") is not None]
    elapsed   = None
    if job.started_at and job.completed_at:
        elapsed = round((job.completed_at - job.started_at).total_seconds())

    return {
        "job_id":        job.job_id,
        "experiment_name": job.experiment.get("name", ""),
        "run_date":      job.started_at.isoformat() if job.started_at else None,
        "completed_at":  job.completed_at.isoformat() if job.completed_at else None,
        "duration_sec":  elapsed,
        "pipeline": {
            "haiku_enabled":           job.pipeline_cfg.get("haiku_enabled"),
            "haiku_validation_enabled":job.pipeline_cfg.get("haiku_validation_enabled"),
            "perplexity_enabled":      job.pipeline_cfg.get("perplexity_enabled"),
            "perplexity_trigger":      job.pipeline_cfg.get("perplexity_trigger"),
            "manual_associations_enabled": job.pipeline_cfg.get("manual_associations_enabled"),
            "salesforce_enrichment_enabled": job.pipeline_cfg.get("salesforce_enrichment_enabled"),
            "legacy_enrichment_enabled":     job.pipeline_cfg.get("legacy_enrichment_enabled"),
            "web_search_fallback_enabled":   job.pipeline_cfg.get("web_search_fallback_enabled"),
            "bing_web_enabled":              job.pipeline_cfg.get("bing_web_enabled"),
            "duckduckgo_web_enabled":        job.pipeline_cfg.get("duckduckgo_web_enabled"),
            "bing_maps_enabled":             job.pipeline_cfg.get("bing_maps_enabled"),
        },
        "job_settings": job.effective_job_settings,
        "runtime": {
            **(job.environment_snapshot or {}),
            "run_started_at": job.started_at.isoformat() if job.started_at else None,
            "run_completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "run_duration_sec": elapsed,
            "last_progress_at": job.last_progress_at.isoformat() if job.last_progress_at else None,
            "last_heartbeat_at": job.last_heartbeat_at.isoformat() if job.last_heartbeat_at else None,
            "max_heartbeat_gap_sec": round(job.max_heartbeat_gap_sec, 1),
            "suspected_sleep_events": job.suspected_sleep_events,
            "stale_timeout_sec": job.effective_job_settings.get("job_stall_timeout_seconds"),
            "stale_detected": job.stale_detected,
            "stale_reason": job.stale_reason,
            "peak_browser_count": job.peak_browser_count,
            "worker_error_count": job.worker_error_count,
            "in_flight_count": getattr(job, "last_in_flight_count", None),
            "active_rows_at_close": active_rows,
            "active_rows_at_close_count": len(active_rows),
        },
        "worker_diagnostics": {
            "worker_error_count": job.worker_error_count,
            "active_rows_at_close": active_rows,
            "event_counts": dict(stage_counts),
            "recent_stage_events": job.stage_events_snapshot(limit=25),
        },
        "input":  {"file_name": job.file_name, "valid_rows": job.total},
        "source": {
            "file_name": job.file_name,
            "file_hash": job.source_metadata.get("source_file_hash"),
            "total_rows_raw": job.source_metadata.get("source_total_rows_raw"),
            "total_valid_rows": job.source_metadata.get("source_total_valid_rows"),
            "valid_row_start": job.source_metadata.get("valid_row_start"),
            "valid_row_end": job.source_metadata.get("valid_row_end"),
            "processed_valid_rows": job.total,
        },
        "coverage": {
            "final_urls_found": job.found_count,
            "without_final_url": job.completed - job.found_count,
            "processed_rows": job.completed,
        },
        "google_maps": {
            "found_listing":           job.maps_found_count,
            "not_found":               job.completed - job.maps_found_count,
            "has_website":             sum(1 for r in results if r.get("gmaps_website")),
            "no_website_on_listing":   statuses.get("no_website", 0),
            "location_match_exact":    sum(1 for r in results if r.get("gmaps_location_match") == "exact"),
            "location_match_partial":  sum(1 for r in results if r.get("gmaps_location_match") == "partial"),
        },
        "confidence": {
            "high":       {"count": tiers.get("High", 0),   "pct": round(tiers.get("High", 0) / completed * 100, 1)},
            "medium":     {"count": tiers.get("Medium", 0), "pct": round(tiers.get("Medium", 0) / completed * 100, 1)},
            "low":        {"count": tiers.get("Low", 0),    "pct": round(tiers.get("Low", 0) / completed * 100, 1)},
            "unresolved": {"count": tiers.get(None, 0),     "pct": round(tiers.get(None, 0) / completed * 100, 1)},
            "avg_score":  round(sum(scores) / len(scores), 1) if scores else None,
            "score_range":[min(scores), max(scores)] if scores else None,
        },
        "status_breakdown": dict(statuses),
        "identity": {
            "accepted": identity.get("accepted", 0),
            "review":   identity.get("review", 0),
            "rejected": identity.get("rejected", 0),
            "no_candidate": identity.get("no_candidate", 0),
        },
        "url_sources": {
            "from_gmaps":      sum(1 for r in results if r.get("url_source") == "gmaps"),
            "from_perplexity": sum(1 for r in results if r.get("url_source") == "perplexity"),
            "from_manual":     sum(1 for r in results if r.get("url_source") == "manual_verified"),
            "from_salesforce": sum(1 for r in results if r.get("url_source") == "salesforce"),
            "from_legacy":     sum(1 for r in results if r.get("url_source") == "legacy_db"),
            "from_web_bing":   sum(1 for r in results if r.get("url_source") == "web_bing"),
            "from_web_duckduckgo": sum(1 for r in results if r.get("url_source") == "web_duckduckgo"),
            "from_bing_maps":  sum(1 for r in results if r.get("url_source") == "bing_maps"),
            "url_changed":     sum(1 for r in results if r.get("url_changed")),
        },
        "search_candidates": {
            "evaluated": sum(r.get("search_candidates_evaluated") or 0 for r in results),
            "gmaps_attempts": sum(r.get("gmaps_attempts") or 0 for r in results),
            "gmaps_retry_recovered": sum(
                1 for r in results
                if r.get("url_source") == "gmaps"
                and r.get("selected_candidate_mode") not in (None, "", "coordinate_company")
            ),
            "bing_maps_recovered": sum(1 for r in results if r.get("url_source") == "bing_maps"),
            "web_search_recovered": sum(
                1 for r in results if r.get("url_source") in {"web_bing", "web_duckduckgo"}
            ),
            "web_candidates": sum(r.get("web_search_candidate_count") or 0 for r in results),
            "web_attempt_rows": sum(1 for r in results if r.get("web_search_attempted")),
            "web_queries": sum(r.get("web_search_query_count") or 0 for r in results),
            "web_diagnostics": sum(r.get("web_search_diagnostic_count") or 0 for r in results),
            "web_errors": sum(r.get("web_search_error_count") or 0 for r in results),
            "web_parsed_results": sum(r.get("web_search_parsed_count") or 0 for r in results),
            "rejected": sum(
                1 for r in results
                for c in r.get("search_candidates", [])
                if (c.get("evaluation") or {}).get("identity_verdict") == "rejected"
            ),
            "rejected_by_reason": dict(Counter(
                ((c.get("evaluation") or {}).get("identity_reason") or "unknown").split(";")[0]
                for r in results
                for c in r.get("search_candidates", [])
                if (c.get("evaluation") or {}).get("identity_verdict") == "rejected"
            )),
        },
        "manual_associations": {
            "enabled": job.pipeline_cfg.get("manual_associations_enabled", True),
            "rows_with_candidates": sum(1 for r in results if r.get("manual_url")),
            "accepted_candidates":  sum(1 for r in results if r.get("manual_identity_verdict") == "accepted"),
            "final_urls":           sum(1 for r in results if r.get("url_source") == "manual_verified"),
        },
        "historical_enrichment": {
            "salesforce_enabled": job.pipeline_cfg.get("salesforce_enrichment_enabled", False),
            "legacy_enabled":     job.pipeline_cfg.get("legacy_enrichment_enabled", False),
            "rows_with_candidates": sum(1 for r in results if (r.get("historical_candidate_count") or 0) > 0),
            "accepted_candidates":  sum(1 for r in results if r.get("historical_identity_verdict") == "accepted"),
            "from_salesforce":      sum(1 for r in results if r.get("url_source") == "salesforce"),
            "from_legacy":          sum(1 for r in results if r.get("url_source") == "legacy_db"),
            "legacy_raw_rows":       sum(r.get("historical_legacy_raw_rows") or 0 for r in results),
            "legacy_rows_with_email":sum(r.get("historical_legacy_rows_with_email") or 0 for r in results),
            "legacy_usable_domains": sum(r.get("historical_legacy_usable_domains") or 0 for r in results),
            "legacy_filtered_domains": sum(r.get("historical_legacy_filtered_domains") or 0 for r in results),
            "errors":               sum(1 for r in results if r.get("historical_errors")),
        },
        "haiku": {
            "calls_initial":  job.haiku_initial_calls,
            "calls_validation": job.haiku_validation_calls,
            "input_tokens":   job.haiku_input_tokens,
            "output_tokens":  job.haiku_output_tokens,
            "cost_usd":       round(job.haiku_cost_usd, 6),
            "truncated_responses": sum(
                1 for r in results if r.get("haiku_initial_stop_reason") == "max_tokens"
                or r.get("haiku_final_stop_reason") == "max_tokens"
            ),
        },
        "perplexity": {
            "calls":          job.perplexity_calls,
            "input_tokens":   job.perplexity_input_tokens,
            "output_tokens":  job.perplexity_output_tokens,
            "cost_usd":       round(job.perplexity_cost_usd, 6),
            "total_citations":sum(len(json.loads(r.get("perplexity_citations") or "[]"))
                                  for r in results),
        },
        "cost": {
            "haiku_usd":           round(job.haiku_cost_usd, 6),
            "perplexity_usd":      round(job.perplexity_cost_usd, 6),
            "total_usd":           round(job.haiku_cost_usd + job.perplexity_cost_usd, 6),
            "cost_per_record":     round((job.haiku_cost_usd + job.perplexity_cost_usd) / completed, 6),
        },
        "performance": {
            "avg_total_latency_ms": round(
                sum(r.get("total_latency_ms") or 0 for r in results) / completed
            ),
            "avg_gmaps_latency_ms": round(
                sum(r.get("gmaps_latency_ms") or 0 for r in results) / completed
            ),
            "avg_bing_maps_latency_ms": round(
                sum(r.get("bing_maps_latency_ms") or 0 for r in results) / completed
            ),
            "avg_web_search_latency_ms": round(
                sum(r.get("web_search_latency_ms") or 0 for r in results) / completed
            ),
            "avg_manual_latency_ms": round(
                sum(r.get("manual_latency_ms") or 0 for r in results) / completed
            ),
            "avg_historical_latency_ms": round(
                sum(r.get("historical_latency_ms") or 0 for r in results) / completed
            ),
            "avg_scoring_latency_ms": round(
                sum(r.get("scoring_latency_ms") or 0 for r in results) / completed
            ),
            "peak_browser_count": job.peak_browser_count,
        },
    }


# ---------------------------------------------------------------------------
# Job model
# ---------------------------------------------------------------------------
class Job:
    def __init__(self, job_id, leads, pipeline_cfg, job_settings,
                 experiment, file_name="", source_metadata=None):
        self.job_id       = job_id
        self.leads        = leads
        self.pipeline_cfg = pipeline_cfg
        self.job_settings = job_settings
        self.experiment   = experiment
        self.file_name    = file_name
        self.source_metadata = source_metadata or {}
        self.status       = Status.PENDING
        self.started_at: Optional[datetime]   = None
        self.completed_at: Optional[datetime] = None
        self.monotonic_started: Optional[float] = None
        self.last_progress_at: Optional[datetime] = None
        self.last_heartbeat_at: Optional[datetime] = None
        self.last_heartbeat_monotonic: Optional[float] = None
        self.max_heartbeat_gap_sec: float = 0.0
        self.suspected_sleep_events: List[dict] = []
        self.stale_detected = False
        self.stale_reason: Optional[str] = None
        self.environment_snapshot: dict = {}
        self.effective_job_settings: dict = {}
        self.peak_browser_count = 0
        self.last_in_flight_count = 0

        self.total     = len(leads)
        self.completed = 0
        self.found_count = 0       # Rows with a selected final URL.
        self.maps_found_count = 0  # Rows where Google Maps returned a listing.
        self.error_count = 0
        self.worker_error_count = 0

        # Per-provider cost tracking
        self.haiku_initial_calls    = 0
        self.haiku_validation_calls = 0
        self.haiku_input_tokens     = 0
        self.haiku_output_tokens    = 0
        self.haiku_cost_usd         = 0.0
        self.perplexity_calls       = 0
        self.perplexity_input_tokens= 0
        self.perplexity_output_tokens=0
        self.perplexity_cost_usd    = 0.0

        self.results:    List[dict] = []
        self.output_file: Optional[str] = None
        self.report:      Optional[dict] = None

        self.cancel_event  = threading.Event()
        self.closed_event  = threading.Event()
        self.event_queue   = queue.Queue(maxsize=1000)
        self._results_lock = threading.Lock()
        self._api_calls:   List[dict] = []
        self._api_lock     = threading.Lock()
        self._active_rows: dict = {}
        self._active_lock = threading.RLock()
        self._stage_events: List[dict] = []
        self._stage_flush_index = 0
        self._stage_lock = threading.Lock()

    def emit(self, event_type: str, data: dict):
        self.last_heartbeat_at = datetime.now(timezone.utc)
        try:
            self.event_queue.put_nowait({"type": event_type, "data": data,
                                          "ts": datetime.now().isoformat()})
        except queue.Full:
            pass

    def mark_progress(self):
        now = datetime.now(timezone.utc)
        self.last_progress_at = now
        self.last_heartbeat_at = now

    def new_row_key(self) -> str:
        return f"{threading.get_ident()}:{time.monotonic_ns()}"

    def _lead_stage_meta(self, lead) -> dict:
        if not lead:
            return {"company": "", "city": "", "state": "", "source_sheet": ""}
        if isinstance(lead, dict):
            return {
                "company": lead.get("company", ""),
                "city": lead.get("city", ""),
                "state": lead.get("state", ""),
                "source_sheet": lead.get("source_sheet", ""),
            }
        return {
            "company": getattr(lead, "company", ""),
            "city": getattr(lead, "city", ""),
            "state": getattr(lead, "state", ""),
            "source_sheet": getattr(lead, "source_sheet", ""),
        }

    def record_stage(self, row_key: Optional[str], event_type: str, stage: str,
                     lead=None, message: str = "", details: Optional[dict] = None,
                     terminal: bool = False, is_active_snapshot: bool = False,
                     emit: bool = True):
        now = datetime.now(timezone.utc)
        mono = time.monotonic()
        meta = self._lead_stage_meta(lead)
        details = details or {}
        elapsed_ms = None

        with self._active_lock:
            if row_key:
                previous = self._active_rows.get(row_key) or {}
                started_mono = previous.get("_started_mono", mono)
                elapsed_ms = int((mono - started_mono) * 1000)
                if terminal:
                    self._active_rows.pop(row_key, None)
                else:
                    self._active_rows[row_key] = {
                        "row_key": row_key,
                        "company": meta["company"],
                        "city": meta["city"],
                        "state": meta["state"],
                        "source_sheet": meta["source_sheet"],
                        "stage": stage,
                        "event_type": event_type,
                        "message": message,
                        "details": details,
                        "worker": threading.current_thread().name,
                        "started_at": previous.get("started_at") or now.isoformat(),
                        "updated_at": now.isoformat(),
                        "elapsed_ms": elapsed_ms,
                        "_started_mono": started_mono,
                    }
            active_rows = self.active_rows_snapshot()

        row = {
            "job_id": self.job_id,
            "event_type": event_type,
            "stage": stage,
            "company": meta["company"],
            "city": meta["city"],
            "state": meta["state"],
            "source_sheet": meta["source_sheet"],
            "worker": threading.current_thread().name,
            "message": message,
            "details_json": json.dumps(details, default=str)[:4000],
            "created_at": now.isoformat(),
            "elapsed_ms": elapsed_ms,
            "is_active_snapshot": bool(is_active_snapshot),
        }
        with self._stage_lock:
            self._stage_events.append(row)

        if emit:
            self.emit("stage", {
                "event": {k: v for k, v in row.items() if k != "details_json"},
                "details": details,
                "active_rows": active_rows,
                "worker_errors": self.worker_error_count,
            })

    def record_worker_error(self, stage: str, exc: Exception, lead=None,
                            row_key: Optional[str] = None):
        with self._results_lock:
            self.worker_error_count += 1
            worker_errors = self.worker_error_count
        message = str(exc)[:250]
        self.record_stage(
            row_key,
            "worker_error",
            stage,
            lead=lead,
            message=message,
            details={"traceback": traceback.format_exc(limit=8)[-3000:]},
            terminal=bool(row_key),
        )
        self.emit("worker_error", {
            "message": message,
            "stage": stage,
            "worker_errors": worker_errors,
            "active_rows": self.active_rows_snapshot(),
        })

    def active_rows_snapshot(self) -> list[dict]:
        with self._active_lock:
            rows = []
            for row in self._active_rows.values():
                clean = {k: v for k, v in row.items() if not k.startswith("_")}
                rows.append(clean)
            rows.sort(key=lambda r: r.get("started_at") or "")
            return rows

    def stage_events_snapshot(self, limit: int = 25) -> list[dict]:
        with self._stage_lock:
            rows = self._stage_events[-limit:] if limit else list(self._stage_events)
            return [dict(r) for r in rows]

    def take_pending_stage_events(self) -> list[dict]:
        with self._stage_lock:
            rows = self._stage_events[self._stage_flush_index:]
            self._stage_flush_index = len(self._stage_events)
            return [dict(r) for r in rows]

    def record_active_snapshot_events(self, event_type: str):
        for row in self.active_rows_snapshot():
            self.record_stage(
                row.get("row_key"),
                event_type,
                row.get("stage") or "unknown",
                lead=row,
                message=row.get("message") or "Row was active when the job stopped.",
                details=row.get("details") or {},
                terminal=False,
                is_active_snapshot=True,
                emit=False,
            )

    def heartbeat(self) -> dict:
        now = datetime.now(timezone.utc)
        mono = time.monotonic()
        if self.last_heartbeat_monotonic is not None:
            gap = mono - self.last_heartbeat_monotonic
            self.max_heartbeat_gap_sec = max(self.max_heartbeat_gap_sec, gap)
            if gap >= 120:
                self.suspected_sleep_events.append({
                    "detected_at": now.isoformat(),
                    "gap_sec": round(gap, 1),
                })
                self.suspected_sleep_events = self.suspected_sleep_events[-10:]
        self.last_heartbeat_monotonic = mono
        self.last_heartbeat_at = now
        return {
            "last_heartbeat_at": now.isoformat(),
            "max_heartbeat_gap_sec": round(self.max_heartbeat_gap_sec, 1),
            "suspected_sleep_events": json.dumps(self.suspected_sleep_events),
        }

    def record_api_call(self, call_meta: dict, stage: str):
        """Thread-safe accumulation of API call metadata."""
        with self._api_lock:
            self._api_calls.append({
                "job_id":        self.job_id,
                "company":       call_meta.get("company", ""),
                "provider":      call_meta.get("provider", ""),
                "stage":         stage,
                "model":         call_meta.get("model", ""),
                "input_tokens":  call_meta.get("input_tokens", 0),
                "output_tokens": call_meta.get("output_tokens", 0),
                "cost_usd":      call_meta.get("cost_usd", 0.0),
                "latency_ms":    call_meta.get("latency_ms", 0),
                "timestamp":     datetime.now(timezone.utc).isoformat(),
                "stop_reason":   call_meta.get("stop_reason", ""),
                "request_id":    call_meta.get("request_id", ""),
                "citations":     json.dumps(call_meta.get("citations", [])),
                "citation_count":call_meta.get("citation_count", 0),
                "raw_response":  call_meta.get("raw_response", "")[:2000],
            })

            # Update job-level provider totals
            provider = call_meta.get("provider", "")
            if provider == "anthropic":
                if stage == "haiku_initial":
                    self.haiku_initial_calls += 1
                else:
                    self.haiku_validation_calls += 1
                self.haiku_input_tokens  += call_meta.get("input_tokens", 0)
                self.haiku_output_tokens += call_meta.get("output_tokens", 0)
                self.haiku_cost_usd      += call_meta.get("cost_usd", 0.0)
            elif provider == "perplexity":
                self.perplexity_calls          += 1
                self.perplexity_input_tokens   += call_meta.get("input_tokens", 0)
                self.perplexity_output_tokens  += call_meta.get("output_tokens", 0)
                self.perplexity_cost_usd       += call_meta.get("cost_usd", 0.0)


# ---------------------------------------------------------------------------
# Global job registry
# ---------------------------------------------------------------------------
_jobs: dict = {}
_jobs_lock  = threading.Lock()


def create_job(leads, pipeline_cfg, job_settings, experiment, file_name="", source_metadata=None) -> str:
    job_id = str(uuid.uuid4())[:8]
    job = Job(job_id, leads, pipeline_cfg, job_settings, experiment, file_name, source_metadata)
    with _jobs_lock:
        _jobs[job_id] = job
    return job_id


def get_job(job_id: str) -> Optional["Job"]:
    return _jobs.get(job_id)


def list_jobs_memory() -> list:
    with _jobs_lock:
        return [
            {"job_id": j.job_id, "status": j.status, "total": j.total,
             "completed": j.completed,
             "cost_usd": round(j.haiku_cost_usd + j.perplexity_cost_usd, 4),
             "created_at": j.started_at.isoformat() if j.started_at else None}
            for j in _jobs.values()
        ]


def cancel_job(job_id: str) -> bool:
    job = get_job(job_id)
    if not job or job.status != Status.RUNNING:
        return False
    job.cancel_event.set()
    job.emit("cancelled", {"message": "Cancellation requested - stopping queued work..."})
    return True


def start_job(job_id: str):
    job = get_job(job_id)
    if not job:
        raise ValueError(f"Job {job_id} not found")
    job.status     = Status.RUNNING
    job.started_at = datetime.now(timezone.utc)
    job.monotonic_started = time.monotonic()
    job.last_progress_at = job.started_at
    job.last_heartbeat_at = job.started_at
    job.last_heartbeat_monotonic = time.monotonic()
    job.environment_snapshot = _environment_snapshot()
    job.effective_job_settings = _effective_job_settings(job.job_settings)

    db.save_job_start(
        job_id          = job_id,
        file_name       = job.file_name,
        valid_rows      = job.total,
        pipeline_config = job.pipeline_cfg,
        experiment      = job.experiment,
        settings_snapshot = settings.pipeline_snapshot(),
        job_settings_snapshot = job.effective_job_settings,
        runtime_snapshot = job.environment_snapshot,
        source_file_hash = job.source_metadata.get("source_file_hash"),
        total_rows_raw = job.source_metadata.get("source_total_rows_raw"),
        source_total_valid_rows = job.source_metadata.get("source_total_valid_rows"),
        valid_row_start = job.source_metadata.get("valid_row_start"),
        valid_row_end = job.source_metadata.get("valid_row_end"),
        created_at      = job.started_at.isoformat(),
    )
    job.emit("status", {"status": "running", "total": job.total})
    threading.Thread(target=_run_pipeline, args=(job,), daemon=True).start()


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------
def _run_pipeline(job: Job):
    cfg = job.pipeline_cfg
    sett= job.job_settings

    haiku_on    = cfg.get("haiku_enabled", settings.haiku_enabled)
    haiku_val_on= cfg.get("haiku_validation_enabled", settings.haiku_validation_enabled)
    perp_on     = cfg.get("perplexity_enabled", settings.perplexity_enabled)
    perp_trigger= cfg.get("perplexity_trigger", settings.perplexity_trigger)
    perp_validate= cfg.get("perplexity_validate", settings.perplexity_validate)
    manual_on   = cfg.get("manual_associations_enabled", settings.manual_associations_enabled)
    sf_enrich_on = cfg.get("salesforce_enrichment_enabled", settings.salesforce_enrichment_enabled)
    legacy_enrich_on = cfg.get("legacy_enrichment_enabled", settings.legacy_enrichment_enabled)
    web_search_on = cfg.get("web_search_fallback_enabled", settings.web_search_fallback_enabled)
    bing_web_on = cfg.get("bing_web_enabled", settings.bing_web_enabled)
    duck_web_on = cfg.get("duckduckgo_web_enabled", settings.duckduckgo_web_enabled)
    bing_maps_on = cfg.get("bing_maps_enabled", settings.bing_maps_enabled)
    manual_association_cache = (
        db.list_manual_associations(include_inactive=False)
        if manual_on else []
    )
    haiku_thresh= sett.get("haiku_confidence_threshold", settings.haiku_confidence_threshold)
    n_scrape    = sett.get("scrape_workers", settings.scrape_workers)
    n_validate  = sett.get("validate_workers", settings.validate_workers)
    web_max_results = sett.get("web_search_max_results", settings.web_search_max_results)
    stale_timeout_sec = int(sett.get(
        "job_stall_timeout_seconds",
        settings.job_stall_timeout_seconds,
    ) or 0)
    last_runtime_db_update = time.monotonic()

    counter       = _Counter()
    scrape_pool   = ThreadPoolExecutor(max_workers=n_scrape,   thread_name_prefix="scrape")
    validate_pool = ThreadPoolExecutor(max_workers=n_validate, thread_name_prefix="validate")
    driver_pool = queue.LifoQueue()
    all_drivers: list = []
    driver_lock = threading.Lock()
    max_pending_scrapes = max(n_scrape * 2, n_scrape)
    lead_iter = iter(job.leads)
    submitted_all_leads = False

    def acquire_driver():
        try:
            return driver_pool.get_nowait()
        except queue.Empty:
            driver = setup_driver()
            with driver_lock:
                all_drivers.append(driver)
                job.peak_browser_count = max(job.peak_browser_count, len(all_drivers))
            return driver

    def release_driver(driver, reusable: bool = True):
        if not driver:
            return
        if reusable and not job.closed_event.is_set():
            try:
                driver_pool.put_nowait(driver)
                return
            except queue.Full:
                pass
        try:
            driver.quit()
        except Exception:
            pass
        with driver_lock:
            try:
                all_drivers.remove(driver)
            except ValueError:
                pass

    def close_all_drivers():
        with driver_lock:
            drivers = list(all_drivers)
            all_drivers.clear()
        while True:
            try:
                driver_pool.get_nowait()
            except queue.Empty:
                break
        for driver in drivers:
            try:
                driver.quit()
            except Exception:
                pass

    def driver_should_recycle(raw: dict) -> bool:
        err = str((raw or {}).get("error") or "").lower()
        return any(token in err for token in (
            "invalid session", "session deleted", "chrome not reachable",
            "disconnected", "no such window", "target window already closed",
        ))

    def flush_stage_events():
        try:
            db.save_stage_events_batch(job.take_pending_stage_events())
        except Exception as exc:
            job.emit("worker_error", {
                "message": f"stage event flush failed: {str(exc)[:200]}",
                "stage": "persist",
                "worker_errors": job.worker_error_count,
                "active_rows": job.active_rows_snapshot(),
            })

    def run_task(fn, *args):
        try:
            fn(*args)
        except Exception as exc:
            lead = None
            row_key = None
            if args:
                first = args[0]
                if isinstance(first, Lead):
                    lead = first
                elif isinstance(first, dict):
                    lead = first.get("lead")
                    row_key = first.get("row_key")
            job.record_worker_error(getattr(fn, "__name__", "task"), exc, lead=lead, row_key=row_key)
            counter.dec()

    def submit_scrape(lead, error_count=0):
        if job.cancel_event.is_set() or job.closed_event.is_set():
            return
        counter.inc()
        try:
            scrape_pool.submit(run_task, _scrape_task, lead, error_count)
        except RuntimeError:
            counter.dec()

    def submit_validate(fn, *args):
        if job.cancel_event.is_set() or job.closed_event.is_set():
            counter.dec()
            return
        try:
            validate_pool.submit(run_task, fn, *args)
        except RuntimeError:
            counter.dec()

    def finalize(ctx: dict):
        if job.closed_event.is_set():
            counter.dec()
            return
        result = _build_result(ctx)
        job.record_stage(
            ctx.get("row_key"),
            "row_finalized",
            "finalize",
            lead=ctx.get("lead"),
            details={
                "status": result.get("status"),
                "final_url": result.get("final_url"),
                "url_source": result.get("url_source"),
                "final_score": result.get("final_confidence_score"),
            },
            terminal=bool(ctx.get("row_key")),
        )
        with job._results_lock:
            job.results.append(result)
            job.completed += 1
            job.mark_progress()
            if result.get("final_url"):
                job.found_count += 1
            if result.get("gmaps_found"):
                job.maps_found_count += 1
            if result.get("status") in ("error", "scrape_error"):
                job.error_count += 1

        total_cost = (ctx.get("haiku_initial_cost", 0) +
                      ctx.get("perplexity_cost", 0) +
                      ctx.get("haiku_final_cost", 0))

        job.emit("progress", {
            "completed": job.completed,
            "total":     job.total,
            "found":     job.found_count,
            "maps_found": job.maps_found_count,
            "errors":    job.error_count,
            "worker_errors": job.worker_error_count,
            "active_rows": job.active_rows_snapshot(),
            "haiku_cost_usd":      round(job.haiku_cost_usd, 4),
            "perplexity_cost_usd": round(job.perplexity_cost_usd, 4),
            "total_cost_usd":      round(job.haiku_cost_usd + job.perplexity_cost_usd, 4),
        })
        job.emit("log", {
            "company":       result.get("company", ""),
            "final_url":     result.get("final_url", ""),
            "url_source":    result.get("url_source", ""),
            "gmaps_score":   result.get("gmaps_confidence_score"),
            "final_score":   result.get("final_confidence_score"),
            "tier":          result.get("confidence_tier", ""),
            "stages_run":    result.get("stages_run", ""),
            "status":        result.get("status", ""),
            "cost_usd":      round(total_cost, 5),
        })
        counter.dec()

    def abandon_current_row(ctx: Optional[dict] = None, row_key: Optional[str] = None,
                            lead: Optional[Lead] = None):
        ctx = ctx or {}
        row_key = row_key or ctx.get("row_key")
        lead = lead or ctx.get("lead")
        if row_key or lead:
            job.record_stage(
                row_key,
                "row_abandoned",
                "cancelled",
                lead=lead,
                message="Row abandoned because the job stopped before finalization.",
                terminal=bool(row_key),
            )
        counter.dec()

    def shutdown_pool(pool):
        try:
            pool.shutdown(wait=False, cancel_futures=True)
        except TypeError:
            pool.shutdown(wait=False)

    def submit_more_scrapes():
        nonlocal submitted_all_leads
        if submitted_all_leads:
            return
        while (
            not job.cancel_event.is_set()
            and not job.closed_event.is_set()
            and counter.value() < max_pending_scrapes
        ):
            try:
                lead = next(lead_iter)
            except StopIteration:
                submitted_all_leads = True
                break
            submit_scrape(lead)

    # ── Stage functions ─────────────────────────────────────────────────

    def _scrape_task(lead: Lead, error_count: int):
        row_key = job.new_row_key()
        row_start = time.time()
        timings = {}
        driver = None
        driver_reusable = True
        job.record_stage(row_key, "row_started", "scrape_start", lead)
        if job.cancel_event.is_set():
            abandon_current_row(row_key=row_key, lead=lead); return

        try:
            job.emit("scraping", {"company": lead.company})
            job.record_stage(row_key, "driver_acquire_start", "driver_pool", lead)
            stage_start = time.time()
            driver = acquire_driver()
            job.record_stage(
                row_key,
                "driver_acquired",
                "driver_pool",
                lead,
                details={"peak_browser_count": job.peak_browser_count},
            )

            def maps_stage_callback(event_type: str, stage: str, details: Optional[dict] = None):
                job.record_stage(row_key, event_type, stage, lead, details=details or {})

            job.record_stage(row_key, "gmaps_start", "gmaps", lead)
            raw = scrape_lead({"company": lead.company, "city": lead.city,
                                "state": lead.state, "country": lead.country,
                                "source_sheet": lead.source_sheet,
                                "gmaps_max_candidates_per_mode": sett.get("gmaps_max_candidates_per_mode", settings.gmaps_max_candidates_per_mode),
                                "gmaps_strong_stop_score": sett.get("gmaps_strong_stop_score", settings.gmaps_strong_stop_score)},
                               driver=driver,
                               stage_callback=maps_stage_callback)
            timings["gmaps_latency_ms"] = int((time.time() - stage_start) * 1000)
            job.record_stage(
                row_key,
                "gmaps_done",
                "gmaps",
                lead,
                details={
                    "attempts": raw.get("gmaps_attempts", 0),
                    "candidates": len(raw.get("gmaps_candidates") or []),
                    "found": bool(raw.get("found")),
                    "has_website": bool(raw.get("website")),
                    "error": raw.get("error"),
                },
            )
            driver_reusable = not driver_should_recycle(raw)
            if job.cancel_event.is_set():
                abandon_current_row(row_key=row_key, lead=lead); return

            # On genuine error (exception in scrape_lead), retry once
            if raw.get("error") and not raw.get("found") and error_count < settings.max_scrape_errors:
                submit_scrape(lead, error_count + 1)
                counter.dec(); return

            addr_parts = parse_gmaps_address(raw.get("address") or "")
            g_score = gmaps_confidence_score(
                gmaps_found    = raw.get("found", False),
                input_name     = lead.company,
                listing_name   = raw.get("gmaps_listing_name"),
                location_match = raw.get("location_match"),
                has_website    = bool(raw.get("website")),
                has_phone      = bool(raw.get("phone")),
                has_address    = bool(raw.get("address")),
            )
            stage_start = time.time()
            search_candidates = _evaluate_search_candidates(
                lead,
                raw.get("gmaps_candidates") or [],
                provider_confidence=g_score,
            )
            timings["scoring_latency_ms"] = timings.get("scoring_latency_ms", 0) + int((time.time() - stage_start) * 1000)

            if bing_maps_on and not _has_promotable_search_candidate(search_candidates):
                job.emit("bing_maps", {"company": lead.company})
                job.record_stage(row_key, "bing_maps_start", "bing_maps", lead)

                def bing_maps_strong_match(candidate: dict) -> bool:
                    stage_start = time.time()
                    evaluated = _evaluate_search_candidates(lead, [candidate])
                    timings["scoring_latency_ms"] = timings.get("scoring_latency_ms", 0) + int((time.time() - stage_start) * 1000)
                    if not evaluated:
                        return False
                    evaluation = evaluated[0].get("evaluation") or {}
                    return (
                        evaluation.get("identity_verdict") == "accepted"
                        and evaluation.get("identity_score", 0) >= sett.get("gmaps_strong_stop_score", settings.gmaps_strong_stop_score)
                    )

                stage_start = time.time()
                bing_maps = lookup_bing_maps_candidates(
                    lead.company,
                    lead.country,
                    city=lead.city,
                    state=lead.state,
                    max_per_mode=int(sett.get("gmaps_max_candidates_per_mode", settings.gmaps_max_candidates_per_mode)),
                    stop_when=bing_maps_strong_match,
                    driver=driver,
                    stage_callback=maps_stage_callback,
                )
                timings["bing_maps_latency_ms"] = int((time.time() - stage_start) * 1000)
                job.record_stage(
                    row_key,
                    "bing_maps_done",
                    "bing_maps",
                    lead,
                    details={
                        "attempts": bing_maps.get("attempts", 0),
                        "candidates": len(bing_maps.get("candidates") or []),
                    },
                )
                stage_start = time.time()
                search_candidates.extend(_evaluate_search_candidates(
                    lead,
                    bing_maps.get("candidates") or [],
                ))
                timings["scoring_latency_ms"] = timings.get("scoring_latency_ms", 0) + int((time.time() - stage_start) * 1000)

            release_driver(driver, reusable=driver_reusable)
            driver = None

            if web_search_on and not _has_promotable_search_candidate(search_candidates):
                job.emit("web_search", {
                    "company": lead.company,
                    "bing": bing_web_on,
                    "duckduckgo": duck_web_on,
                })
                job.record_stage(row_key, "web_search_start", "web_search", lead)
                stage_start = time.time()
                web_candidates = lookup_web_candidates(
                    lead.company,
                    lead.city,
                    lead.state,
                    use_bing=bing_web_on,
                    use_duckduckgo=duck_web_on,
                    max_results=web_max_results,
                )
                timings["web_search_latency_ms"] = int((time.time() - stage_start) * 1000)
                job.record_stage(
                    row_key,
                    "web_search_done",
                    "web_search",
                    lead,
                    details={"candidates": len(web_candidates or [])},
                )
                stage_start = time.time()
                search_candidates.extend(_evaluate_search_candidates(lead, web_candidates))
                timings["scoring_latency_ms"] = timings.get("scoring_latency_ms", 0) + int((time.time() - stage_start) * 1000)
        finally:
            if driver:
                release_driver(driver, reusable=driver_reusable)

        best_maps_for_context = _best_search_candidate(search_candidates, {"gmaps"})
        if best_maps_for_context and _candidate_is_promotable(
            best_maps_for_context.get("evaluation") or {}, best_maps_for_context
        ):
            best_raw = best_maps_for_context.get("raw") or {}
            raw = dict(raw)
            raw.update({
                "website": best_raw.get("website") or best_maps_for_context.get("url"),
                "gmaps_listing_name": best_raw.get("gmaps_listing_name") or best_maps_for_context.get("title"),
                "address": best_raw.get("address") or best_maps_for_context.get("address_or_snippet"),
                "phone": best_raw.get("phone") or best_maps_for_context.get("phone"),
                "gmaps_url": best_raw.get("gmaps_url") or best_maps_for_context.get("maps_url"),
                "location_match": best_raw.get("location_match") or best_maps_for_context.get("location_match"),
                "found": best_raw.get("found", raw.get("found")),
            })
            addr_parts = parse_gmaps_address(raw.get("address") or "")
            g_score = gmaps_confidence_score(
                gmaps_found    = raw.get("found", False),
                input_name     = lead.company,
                listing_name   = raw.get("gmaps_listing_name"),
                location_match = raw.get("location_match"),
                has_website    = bool(raw.get("website")),
                has_phone      = bool(raw.get("phone")),
                has_address    = bool(raw.get("address")),
            )

        ctx = {
            "lead":         lead,
            "scrape":       raw,
            "addr_parts":   addr_parts,
            "gmaps_score":  g_score,
            "search_candidates": search_candidates,
            "stages_run":   ["gmaps"],
            "start_time":   row_start,
            "timings":      timings,
            "row_key":      row_key,
        }
        if any(c.get("source") == "bing_maps" for c in search_candidates):
            ctx["stages_run"].append("bing_maps")
        if any(c.get("source") in {"web_bing", "web_duckduckgo"} for c in search_candidates):
            ctx["stages_run"].append("web_search")
        if manual_on:
            job.emit("manual_association", {"company": lead.company})
            job.record_stage(row_key, "manual_start", "manual_association", lead)
            stage_start = time.time()
            manual = lookup_manual_associations(
                company=lead.company,
                city=lead.city,
                state=lead.state,
                country=lead.country,
                associations=manual_association_cache,
            )
            timings["manual_latency_ms"] = int((time.time() - stage_start) * 1000)
            job.record_stage(
                row_key,
                "manual_done",
                "manual_association",
                lead,
                details={"candidates_found": bool(manual.get("candidates_found"))},
            )
            ctx["manual"] = manual
            if manual.get("candidates_found"):
                ctx["stages_run"].append("manual")
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return

        if sf_enrich_on or legacy_enrich_on:
            job.emit("historical_enrichment", {
                "company": lead.company,
                "salesforce": sf_enrich_on,
                "legacy": legacy_enrich_on,
            })
            job.record_stage(
                row_key,
                "historical_start",
                "historical_enrichment",
                lead,
                details={"salesforce": sf_enrich_on, "legacy": legacy_enrich_on},
            )
            stage_start = time.time()
            hist = lookup_historical_candidates(
                company=lead.company,
                city=lead.city,
                state=lead.state,
                country=lead.country,
                use_salesforce=sf_enrich_on,
                use_legacy=legacy_enrich_on,
                source_objects=settings.salesforce_enrichment_objects,
            )
            timings["historical_latency_ms"] = int((time.time() - stage_start) * 1000)
            job.record_stage(
                row_key,
                "historical_done",
                "historical_enrichment",
                lead,
                details={
                    "candidates": hist.get("candidate_count"),
                    "errors": hist.get("errors"),
                },
            )
            ctx["historical"] = hist
            ctx["stages_run"].append("historical")
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return

        has_url = bool(raw.get("website")) and _has_promotable_search_candidate(
            [c for c in search_candidates if c.get("source") == "gmaps"]
        )
        bing_maps_has_url = _has_promotable_search_candidate(
            [c for c in search_candidates if c.get("source") == "bing_maps"]
        )
        ctx["gmaps_context_usable"] = has_url or not raw.get("website")
        manual_best = (ctx.get("manual") or {}).get("best") or {}
        manual_eval = manual_best.get("evaluation") or {}
        manual_accepted = manual_eval.get("identity_verdict") == "accepted"
        hist_best = (ctx.get("historical") or {}).get("best") or {}
        hist_eval = hist_best.get("evaluation") or {}
        historical_accepted = (
            hist_eval.get("identity_verdict") == "accepted"
            and hist_eval.get("identity_score", 0) >= settings.historical_enrichment_min_score
        )

        # Route to next stage
        if manual_accepted and not (perp_on and perp_trigger == "always"):
            finalize(ctx)
        elif historical_accepted and not (perp_on and perp_trigger == "always"):
            finalize(ctx)
        elif bing_maps_has_url and not (perp_on and perp_trigger == "always"):
            finalize(ctx)
        elif haiku_on and has_url:
            submit_validate(_haiku_initial_task, ctx)
        elif perp_on and _should_perplexity_run(perp_trigger, has_url, g_score, haiku_thresh):
            submit_validate(_perplexity_task, ctx, g_score)
        else:
            finalize(ctx)

    def _haiku_initial_task(ctx: dict):
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return

        lead = ctx["lead"]
        url  = ctx["scrape"].get("website", "")
        job.emit("haiku_initial", {"company": lead.company, "url": url})
        job.record_stage(ctx.get("row_key"), "haiku_initial_start", "haiku_initial", lead)

        snippet, fetch_err = fetch_page_snippet(url)
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return
        if not snippet:
            ctx["haiku_initial_error"] = fetch_err
        else:
            result_meta = validate_with_haiku(
                company       = lead.company,
                city          = lead.city,
                state         = lead.state,
                page_snippet  = snippet,
                gmaps_address = ctx["scrape"].get("address"),
                gmaps_zip     = ctx["addr_parts"].get("gmaps_zip"),
            )
            if job.cancel_event.is_set():
                abandon_current_row(ctx); return
            result_meta["company"] = lead.company
            job.record_api_call(result_meta, "haiku_initial")

            signals = result_meta.get("signals", {})
            h_score = haiku_confidence_score(ctx["gmaps_score"], signals)

            ctx["haiku_initial"] = {
                "signals":     signals,
                "score":       h_score,
                "match":       signals.get("name_match") in ("YES", "PARTIAL"),
                "stop_reason": result_meta.get("stop_reason"),
                "latency_ms":  result_meta.get("latency_ms"),
                "cost":        result_meta.get("cost_usd", 0.0),
            }
            ctx["haiku_initial_cost"] = result_meta.get("cost_usd", 0.0)
            ctx["stages_run"].append("haiku_initial")
            job.record_stage(
                ctx.get("row_key"),
                "haiku_initial_done",
                "haiku_initial",
                lead,
                details={"score": h_score, "latency_ms": result_meta.get("latency_ms")},
            )

            # If Haiku is confident, done
            if h_score >= haiku_thresh and not job.cancel_event.is_set():
                finalize(ctx); return

        # Pass to Perplexity if enabled
        if perp_on and not job.cancel_event.is_set():
            submit_validate(_perplexity_task, ctx,
                            ctx.get("haiku_initial", {}).get("score"))
        else:
            finalize(ctx)

    def _perplexity_task(ctx: dict, haiku_initial_score: Optional[int]):
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return

        lead       = ctx["lead"]
        scrape     = ctx["scrape"]
        addr       = ctx["addr_parts"]
        has_url    = bool(scrape.get("website")) and _has_promotable_search_candidate(
            [c for c in (ctx.get("search_candidates") or []) if c.get("source") == "gmaps"]
        )
        maps_context_usable = ctx.get("gmaps_context_usable", True)

        trigger_score = haiku_initial_score if haiku_initial_score is not None else ctx.get("gmaps_score")
        if not _should_perplexity_run(perp_trigger, has_url, trigger_score, haiku_thresh):
            finalize(ctx); return

        mode = "validate" if (perp_validate and has_url) else "find"
        job.emit("perplexity", {"company": lead.company, "mode": mode})
        job.record_stage(ctx.get("row_key"), "perplexity_start", "perplexity", lead, details={"mode": mode})

        result_meta = call_perplexity(
            company      = lead.company,
            city         = lead.city,
            state        = lead.state,
            gmaps_street = addr.get("gmaps_street") if maps_context_usable else None,
            gmaps_zip    = addr.get("gmaps_zip") if maps_context_usable else None,
            gmaps_phone  = scrape.get("phone") if maps_context_usable else None,
            existing_url = scrape.get("website") if mode == "validate" else None,
            mode         = mode,
        )
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return
        result_meta["company"] = lead.company
        job.record_api_call(result_meta, "perplexity")

        parsed = result_meta.get("parsed", {})
        p_url  = parsed.get("url")

        ctx["perplexity"] = {
            "url":         p_url,
            "reason":      parsed.get("reason"),
            "official_name":     parsed.get("official_name"),
            "evidence_location": parsed.get("evidence_location"),
            "evidence_url":      parsed.get("evidence_url"),
            "company_match":     parsed.get("company_match"),
            "location_match":    parsed.get("location_match"),
            "is_official":       parsed.get("is_official"),
            "reject_reason":     parsed.get("reject_reason"),
            "citations":   result_meta.get("citations", []),
            "latency_ms":  result_meta.get("latency_ms"),
            "cost":        result_meta.get("cost_usd", 0.0),
            "mode":        mode,
        }
        ctx["perplexity_cost"] = result_meta.get("cost_usd", 0.0)
        ctx["stages_run"].append("perplexity")
        job.record_stage(
            ctx.get("row_key"),
            "perplexity_done",
            "perplexity",
            lead,
            details={"mode": mode, "has_url": bool(p_url), "latency_ms": result_meta.get("latency_ms")},
        )

        # Second Haiku pass to validate Perplexity's URL
        if haiku_val_on and p_url and not job.cancel_event.is_set():
            submit_validate(_haiku_validation_task, ctx)
        else:
            finalize(ctx)

    def _haiku_validation_task(ctx: dict):
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return

        lead  = ctx["lead"]
        p_url = ctx["perplexity"]["url"]
        job.emit("haiku_validation", {"company": lead.company, "url": p_url})
        job.record_stage(ctx.get("row_key"), "haiku_validation_start", "haiku_validation", lead)

        snippet, fetch_err = fetch_page_snippet(p_url)
        if job.cancel_event.is_set():
            abandon_current_row(ctx); return
        if snippet:
            result_meta = validate_with_haiku(
                company       = lead.company,
                city          = lead.city,
                state         = lead.state,
                page_snippet  = snippet,
                gmaps_address = ctx["scrape"].get("address"),
                gmaps_zip     = ctx["addr_parts"].get("gmaps_zip"),
            )
            if job.cancel_event.is_set():
                abandon_current_row(ctx); return
            result_meta["company"] = lead.company
            job.record_api_call(result_meta, "haiku_validation")

            signals = result_meta.get("signals", {})
            h_score = haiku_confidence_score(ctx["gmaps_score"], signals)

            ctx["haiku_final"] = {
                "signals":     signals,
                "score":       h_score,
                "match":       signals.get("name_match") in ("YES", "PARTIAL"),
                "stop_reason": result_meta.get("stop_reason"),
                "latency_ms":  result_meta.get("latency_ms"),
                "cost":        result_meta.get("cost_usd", 0.0),
            }
            ctx["haiku_final_cost"] = result_meta.get("cost_usd", 0.0)
            ctx["stages_run"].append("haiku_validation")
            job.record_stage(
                ctx.get("row_key"),
                "haiku_validation_done",
                "haiku_validation",
                lead,
                details={"score": h_score, "latency_ms": result_meta.get("latency_ms")},
            )

        finalize(ctx)

    # ── Kick off ────────────────────────────────────────────────────────
    submit_more_scrapes()

    while True:
        no_in_flight = counter.wait(timeout=0.5)
        in_flight = counter.value()
        job.last_in_flight_count = in_flight
        if submitted_all_leads and no_in_flight:
            shutdown_pool(scrape_pool)
            shutdown_pool(validate_pool)
            break

        submit_more_scrapes()
        now_mono = time.monotonic()
        if now_mono - last_runtime_db_update >= 30:
            heartbeat = job.heartbeat()
            db.update_job_runtime(job.job_id, {
                **heartbeat,
                "completed_rows": job.completed,
                "found_on_maps": job.maps_found_count,
                "error_count": job.error_count,
                "worker_error_count": job.worker_error_count,
                "active_worker_snapshot": json.dumps(job.active_rows_snapshot(), default=str),
                "peak_browser_count": job.peak_browser_count,
            })
            flush_stage_events()
            last_runtime_db_update = now_mono

        if stale_timeout_sec > 0 and job.last_progress_at:
            idle_sec = (datetime.now(timezone.utc) - job.last_progress_at).total_seconds()
            active_rows = job.active_rows_snapshot()
            if in_flight > 0 and not active_rows and idle_sec >= 120:
                job.stale_detected = True
                job.stale_reason = (
                    f"scheduler_counter_stuck_without_active_rows_for_{int(idle_sec)}s_"
                    f"with_{in_flight}_in_flight"
                )
                job.cancel_event.set()
                job.record_stage(
                    None,
                    "scheduler_stale_no_active_rows",
                    "scheduler",
                    message=job.stale_reason,
                    details={
                        "idle_sec": int(idle_sec),
                        "in_flight": in_flight,
                        "completed": job.completed,
                        "total": job.total,
                    },
                )
                job.emit("stale", {
                    "message": "Scheduler stalled with no active rows; saving a partial run.",
                    "idle_sec": int(idle_sec),
                    "in_flight": in_flight,
                    "stale_timeout_sec": 120,
                })
                shutdown_pool(scrape_pool)
                shutdown_pool(validate_pool)
                break
            if in_flight > 0 and idle_sec >= stale_timeout_sec:
                job.stale_detected = True
                job.stale_reason = (
                    f"no_completed_rows_for_{int(idle_sec)}s_with_{in_flight}_in_flight"
                )
                job.cancel_event.set()
                job.emit("stale", {
                    "message": "No rows completed within the stale timeout; saving a partial run.",
                    "idle_sec": int(idle_sec),
                    "in_flight": in_flight,
                    "stale_timeout_sec": stale_timeout_sec,
                })
                shutdown_pool(scrape_pool)
                shutdown_pool(validate_pool)
                break

        if job.cancel_event.is_set():
            job.emit("cancelled", {"message": "Cancellation requested - queued work stopped."})
            shutdown_pool(scrape_pool)
            shutdown_pool(validate_pool)
            break

    job.closed_event.set()
    if job.cancel_event.is_set() or job.stale_detected:
        job.record_active_snapshot_events("row_active_on_cancel")
    close_all_drivers()

    job.completed_at  = datetime.now(timezone.utc)
    job.heartbeat()
    final_status      = Status.CANCELLED if job.cancel_event.is_set() else Status.DONE

    try:
        # Save all API calls to DB
        db.save_api_calls_batch(job._api_calls)
        flush_stage_events()

        # Save individual results first so candidate rows can point at result_id.
        job.record_stage(None, "persist_start", "persist", details={
            "results": len(job.results),
            "api_calls": len(job._api_calls),
        })
        for idx, r in enumerate(job.results, 1):
            candidate_count = len(r.get("search_candidates", []) or [])
            job.record_stage(None, "result_save_start", "persist", lead=r, details={
                "index": idx,
                "results": len(job.results),
                "candidate_count": candidate_count,
            })
            candidate_rows = [
                _search_candidate_to_db_row(job.job_id, 0, r, candidate)
                for candidate in r.get("search_candidates", [])
            ]
            candidate_rows = [_cap_search_candidate_db_row(row) for row in candidate_rows]
            if candidate_rows:
                job.record_stage(None, "candidate_save_start", "persist", lead=r, details={
                    "index": idx,
                    "candidate_count": len(candidate_rows),
                })

            candidates_saved = bool(candidate_rows)
            try:
                result_id = db.save_result_with_candidates(_result_to_db_row(job.job_id, r), candidate_rows)
            except Exception as exc:
                candidates_saved = False
                job.record_stage(None, "result_save_with_candidates_failed", "persist", lead=r,
                                 message=str(exc)[:250],
                                 details={
                                     "index": idx,
                                     "candidate_count": len(candidate_rows),
                                 })
                with job._results_lock:
                    job.worker_error_count += 1

                try:
                    result_id = db.save_result(_result_to_db_row(job.job_id, r))
                    job.record_stage(None, "result_save_fallback_done", "persist", lead=r, details={
                        "index": idx,
                        "result_id": result_id,
                        "candidate_count": len(candidate_rows),
                    })
                    if candidate_rows:
                        try:
                            candidate_rows = [dict(row, result_id=result_id) for row in candidate_rows]
                            db.save_search_candidates_batch(candidate_rows)
                            candidates_saved = True
                            job.record_stage(None, "candidate_save_fallback_done", "persist", lead=r, details={
                                "index": idx,
                                "result_id": result_id,
                                "candidate_count": len(candidate_rows),
                            })
                        except Exception as candidate_exc:
                            job.record_stage(None, "candidate_save_failed", "persist", lead=r,
                                             message=str(candidate_exc)[:250],
                                             details={
                                                 "index": idx,
                                                 "result_id": result_id,
                                                 "candidate_count": len(candidate_rows),
                                             })
                            with job._results_lock:
                                job.worker_error_count += 1
                except Exception as result_exc:
                    job.record_stage(None, "result_save_fallback_failed", "persist", lead=r,
                                     message=str(result_exc)[:250],
                                     details={
                                         "index": idx,
                                         "candidate_count": len(candidate_rows),
                                     })
                    with job._results_lock:
                        job.worker_error_count += 1
                    continue

            job.record_stage(None, "result_save_done", "persist", lead=r, details={
                "index": idx,
                "result_id": result_id,
            })
            if candidates_saved:
                job.record_stage(None, "candidate_save_done", "persist", lead=r, details={
                    "index": idx,
                    "result_id": result_id,
                    "candidate_count": len(candidate_rows),
                })

        job.record_stage(None, "report_build_start", "persist")
        job.report  = build_report(job)
        job.record_stage(None, "workbook_write_start", "persist")
        out_path   = _write_output(job)
        job.record_stage(None, "workbook_write_done", "persist", details={"output_file": out_path})
        job.output_file = out_path
        total_cost  = job.haiku_cost_usd + job.perplexity_cost_usd

        job.record_stage(None, "job_complete_save_start", "persist")
        db.save_job_complete(job.job_id, {
            "status":          final_status,
            "completed_at":    job.completed_at.isoformat(),
            "completed_rows":  job.completed,
            "run_duration_sec": round((job.completed_at - job.started_at).total_seconds()) if job.started_at else None,
            "last_heartbeat_at": job.last_heartbeat_at.isoformat() if job.last_heartbeat_at else None,
            "max_heartbeat_gap_sec": round(job.max_heartbeat_gap_sec, 1),
            "suspected_sleep_events": json.dumps(job.suspected_sleep_events),
            "stale_timeout_sec": stale_timeout_sec,
            "stale_detected": job.stale_detected,
            "stale_reason": job.stale_reason,
            "peak_browser_count": job.peak_browser_count,
            "worker_error_count": job.worker_error_count,
            "active_worker_snapshot": json.dumps(job.active_rows_snapshot(), default=str),
            "found_on_maps":   job.maps_found_count,
            "error_count":     job.error_count,
            "has_website":     sum(1 for r in job.results if r.get("gmaps_website")),
            "tier_high":       sum(1 for r in job.results if r.get("confidence_tier") == "High"),
            "tier_medium":     sum(1 for r in job.results if r.get("confidence_tier") == "Medium"),
            "tier_low":        sum(1 for r in job.results if r.get("confidence_tier") == "Low"),
            "tier_none":       sum(1 for r in job.results if r.get("confidence_tier") is None),
            "status_confirmed":sum(1 for r in job.results if r.get("status") == "confirmed"),
            "status_medium_conf": sum(1 for r in job.results if r.get("status") == "medium_confidence"),
            "status_no_website":  sum(1 for r in job.results if r.get("status") == "no_website"),
            "status_not_found":   sum(1 for r in job.results if r.get("status") == "not_found"),
            "status_manual_verified": sum(1 for r in job.results if r.get("status") == "manual_verified"),
            "manual_rows_with_candidates": sum(1 for r in job.results if r.get("manual_url")),
            "manual_accepted_candidates":  sum(1 for r in job.results if r.get("manual_identity_verdict") == "accepted"),
            "manual_final_urls":           sum(1 for r in job.results if r.get("url_source") == "manual_verified"),
            "status_historical_found": sum(1 for r in job.results if r.get("status") == "historical_found"),
            "historical_rows_with_candidates": sum(1 for r in job.results if (r.get("historical_candidate_count") or 0) > 0),
            "historical_accepted_candidates":  sum(1 for r in job.results if r.get("historical_identity_verdict") == "accepted"),
            "historical_from_salesforce":      sum(1 for r in job.results if r.get("url_source") == "salesforce"),
            "historical_from_legacy":          sum(1 for r in job.results if r.get("url_source") == "legacy_db"),
            "historical_error_rows":           sum(1 for r in job.results if r.get("historical_errors")),
            "search_candidates_evaluated": sum(r.get("search_candidates_evaluated") or 0 for r in job.results),
            "gmaps_attempts":              sum(r.get("gmaps_attempts") or 0 for r in job.results),
            "gmaps_retry_recovered":       sum(
                1 for r in job.results
                if r.get("url_source") == "gmaps"
                and r.get("selected_candidate_mode") not in (None, "", "coordinate_company")
            ),
            "bing_maps_recovered":         sum(1 for r in job.results if r.get("url_source") == "bing_maps"),
            "web_search_recovered":        sum(1 for r in job.results if r.get("url_source") in {"web_bing", "web_duckduckgo"}),
            "web_search_attempt_rows":     sum(1 for r in job.results if r.get("web_search_attempted")),
            "web_search_query_count":      sum(r.get("web_search_query_count") or 0 for r in job.results),
            "web_search_diagnostic_count": sum(r.get("web_search_diagnostic_count") or 0 for r in job.results),
            "web_search_error_count":      sum(r.get("web_search_error_count") or 0 for r in job.results),
            "web_search_parsed_count":     sum(r.get("web_search_parsed_count") or 0 for r in job.results),
            "haiku_calls":     job.haiku_initial_calls + job.haiku_validation_calls,
            "haiku_input_tokens":  job.haiku_input_tokens,
            "haiku_output_tokens": job.haiku_output_tokens,
            "haiku_cost_usd":      job.haiku_cost_usd,
            "perplexity_calls":    job.perplexity_calls,
            "perplexity_input_tokens":  job.perplexity_input_tokens,
            "perplexity_output_tokens": job.perplexity_output_tokens,
            "perplexity_cost_usd":      job.perplexity_cost_usd,
            "cost_usd":        total_cost,
            "avg_confidence_score": (
                round(sum(r["final_confidence_score"] for r in job.results
                          if r.get("final_confidence_score") is not None)
                      / max(1, sum(1 for r in job.results if r.get("final_confidence_score") is not None)), 1)
            ),
            "output_file":  out_path,
            "report_json":  json.dumps(job.report),
        })
        job.record_stage(None, "job_complete_save_done", "persist")
        flush_stage_events()

        db.save_api_calls_batch([])   # batch already saved above

        job.emit("done", {
            "output_file":  out_path,
            "total":        job.total,
            "completed":    job.completed,
            "found":        job.found_count,
            "maps_found":   job.maps_found_count,
            "worker_errors": job.worker_error_count,
            "active_rows": job.active_rows_snapshot(),
            "haiku_cost_usd":      round(job.haiku_cost_usd, 6),
            "perplexity_cost_usd": round(job.perplexity_cost_usd, 6),
            "total_cost_usd":      round(total_cost, 6),
            "cancelled":    job.cancel_event.is_set(),
            "report":       job.report,
        })
    except Exception as exc:
        final_status = Status.ERROR
        job.status = Status.ERROR
        job.completed_at = datetime.now(timezone.utc)
        job.record_stage(None, "persist_failed", "persist", message=str(exc)[:250],
                         details={"error": str(exc)[:1000]})
        with job._results_lock:
            job.worker_error_count += 1
        flush_stage_events()
        try:
            db.save_job_complete(job.job_id, {
                "status": Status.ERROR,
                "completed_at": job.completed_at.isoformat(),
                "completed_rows": job.completed,
                "run_duration_sec": round((job.completed_at - job.started_at).total_seconds()) if job.started_at else None,
                "last_heartbeat_at": job.last_heartbeat_at.isoformat() if job.last_heartbeat_at else None,
                "max_heartbeat_gap_sec": round(job.max_heartbeat_gap_sec, 1),
                "suspected_sleep_events": json.dumps(job.suspected_sleep_events),
                "stale_timeout_sec": stale_timeout_sec,
                "stale_detected": True,
                "stale_reason": f"completion_failed: {str(exc)[:200]}",
                "peak_browser_count": job.peak_browser_count,
                "worker_error_count": job.worker_error_count,
                "active_worker_snapshot": json.dumps(job.active_rows_snapshot(), default=str),
                "found_on_maps": job.maps_found_count,
                "error_count": job.error_count,
                "has_website": sum(1 for r in job.results if r.get("gmaps_website")),
                "tier_high": sum(1 for r in job.results if r.get("confidence_tier") == "High"),
                "tier_medium": sum(1 for r in job.results if r.get("confidence_tier") == "Medium"),
                "tier_low": sum(1 for r in job.results if r.get("confidence_tier") == "Low"),
                "tier_none": sum(1 for r in job.results if r.get("confidence_tier") is None),
                "perplexity_calls": job.perplexity_calls,
                "perplexity_input_tokens": job.perplexity_input_tokens,
                "perplexity_output_tokens": job.perplexity_output_tokens,
                "perplexity_cost_usd": job.perplexity_cost_usd,
                "haiku_calls": job.haiku_initial_calls + job.haiku_validation_calls,
                "haiku_input_tokens": job.haiku_input_tokens,
                "haiku_output_tokens": job.haiku_output_tokens,
                "haiku_cost_usd": job.haiku_cost_usd,
                "cost_usd": job.haiku_cost_usd + job.perplexity_cost_usd,
                "output_file": job.output_file,
                "report_json": json.dumps(job.report) if job.report else None,
            })
        except Exception:
            pass
        job.emit("error", {
            "message": str(exc),
            "terminal": True,
            "total": job.total,
            "completed": job.completed,
            "found": job.found_count,
            "maps_found": job.maps_found_count,
            "errors": job.error_count,
            "worker_errors": job.worker_error_count,
            "active_rows": job.active_rows_snapshot(),
        })

    job.status = final_status


# ---------------------------------------------------------------------------
# Perplexity trigger logic
# ---------------------------------------------------------------------------
def _should_perplexity_run(trigger: str, has_url: bool,
                            haiku_score: Optional[int],
                            haiku_thresh: int) -> bool:
    if trigger == "always":
        return True
    if trigger == "no_url_only":
        return not has_url
    # default: no_url_or_low_confidence
    if not has_url:
        return True
    if haiku_score is not None and haiku_score < haiku_thresh:
        return True
    return False


def _evaluate_search_candidates(lead: Lead, candidates: list[dict],
                                provider_confidence: Optional[int] = None) -> list[dict]:
    evaluated = []
    for candidate in candidates:
        source = candidate.get("source")
        if candidate.get("diagnostic"):
            item = dict(candidate)
            if candidate.get("url"):
                reason = candidate.get("error") or "diagnostic_link_not_eligible"
                item["evaluation"] = {
                    "candidate_url": candidate.get("url"),
                    "identity_score": 0,
                    "identity_verdict": "rejected",
                    "identity_reason": f"Diagnostic link rejected: {reason}",
                    "company_match_score": 0,
                    "domain_match_score": 0,
                    "location_match_level": "unknown",
                    "provider_confidence": None,
                }
            else:
                item["evaluation"] = score_candidate_url(
                    source_company=lead.company,
                    source_city=lead.city,
                    source_state=lead.state,
                    source_country=lead.country,
                    candidate_url=None,
                    candidate_source=source,
                )
            evaluated.append(item)
            continue
        map_like = source in {"gmaps", "bing_maps"}
        evaluation = score_candidate_url(
            source_company=lead.company,
            source_city=lead.city,
            source_state=lead.state,
            source_country=lead.country,
            candidate_url=candidate.get("url"),
            candidate_source=source,
            candidate_name=_candidate_name_for_scoring(candidate),
            candidate_address=candidate.get("address_or_snippet") if map_like else None,
            gmaps_location_match=candidate.get("location_match") if map_like else None,
            evidence_location=candidate.get("address_or_snippet") if not map_like else None,
            provider_confidence=provider_confidence if source == "gmaps" else None,
        )
        evaluation = _apply_retry_gmaps_identity_gate(lead, candidate, evaluation)
        evaluation = _apply_bing_maps_identity_gate(lead, candidate, evaluation)
        evaluation = _apply_web_search_identity_gate(lead, candidate, evaluation)
        item = dict(candidate)
        item["evaluation"] = evaluation
        evaluated.append(item)
    return evaluated


def _candidate_name_for_scoring(candidate: dict) -> str | None:
    if candidate.get("source") in {"web_bing", "web_duckduckgo"}:
        return " ".join(
            str(v).strip()
            for v in [candidate.get("title"), candidate.get("address_or_snippet")]
            if v
        ) or None
    return candidate.get("title")


def _apply_retry_gmaps_identity_gate(lead: Lead, candidate: dict, evaluation: dict) -> dict:
    """
    Retry-mode Maps searches are useful for discovery, but noisy. Require
    distinctive identity evidence before a retry-mode Maps candidate can be
    treated like a final website.
    """
    if candidate.get("source") != "gmaps" or candidate.get("mode") == "coordinate_company":
        return evaluation
    if not candidate.get("url"):
        return evaluation

    company_score = evaluation.get("company_match_score") or 0
    domain_score = evaluation.get("domain_match_score") or 0
    location_level = evaluation.get("location_match_level")
    overlap = business_identity_overlap(lead.company, candidate.get("title"))
    directory_or_provider = "directory_or_social_url" in (evaluation.get("identity_reason") or "").lower()
    has_map_location = location_level in {"exact", "nearby"}
    strong_identity = (
        not directory_or_provider
        and (
            (
                overlap is not None
                and overlap >= 0.75
                and (
                    (company_score >= 95 and domain_score >= 55)
                    or (company_score >= 88 and domain_score >= 75)
                )
            )
            or (
                has_map_location
                and company_score >= 90
                and domain_score >= 55
            )
            or (
                has_map_location
                and company_score >= 85
                and domain_score >= 85
            )
        )
    )

    reason = evaluation.get("identity_reason") or ""
    overlap_label = "none" if overlap is None else round(overlap, 2)
    extra_reason = f"retry_mode; business_overlap={overlap_label}"

    if strong_identity:
        out = dict(evaluation)
        out["identity_reason"] = "; ".join(x for x in [reason, extra_reason, "retry_identity_strong"] if x)
        return out

    out = dict(evaluation)
    cap = 49 if (
        directory_or_provider
        or location_level == "contradiction"
        or company_score < 85
        or domain_score < 55
        or (not has_map_location and (overlap is None or overlap < 0.75))
    ) else 84
    out["identity_score"] = min(out.get("identity_score", 0), cap)
    out["identity_verdict"] = "rejected" if out["identity_score"] < 50 else "review"
    out["identity_reason"] = "; ".join(
        x for x in [reason, extra_reason, "retry_mode_identity_gate"] if x
    )
    return out


def _apply_web_search_identity_gate(lead: Lead, candidate: dict, evaluation: dict) -> dict:
    """
    Browser-search results are noisy and often broad-term matches. Keep weak
    web hits as review candidates, but require stronger identity evidence
    before a web candidate can be automatically accepted.
    """
    if candidate.get("source") not in {"web_bing", "web_duckduckgo"}:
        return evaluation
    if candidate.get("diagnostic") or not candidate.get("url"):
        return evaluation

    company_score = evaluation.get("company_match_score") or 0
    domain_score = evaluation.get("domain_match_score") or 0
    location_level = evaluation.get("location_match_level")
    evidence_text = " ".join(
        str(v).strip()
        for v in [candidate.get("title"), candidate.get("address_or_snippet")]
        if v
    )
    overlap = business_identity_overlap(lead.company, evidence_text)
    has_location_evidence = location_level in {"exact", "nearby", "state_only"}
    strong_identity = (
        company_score >= 85
        and domain_score >= 85
        and location_level != "contradiction"
        and (has_location_evidence or (overlap is not None and overlap >= 0.75))
    )

    reason = evaluation.get("identity_reason") or ""
    overlap_label = "none" if overlap is None else round(overlap, 2)
    extra_reason = f"web_search_gate; business_overlap={overlap_label}"

    if strong_identity:
        out = dict(evaluation)
        out["identity_reason"] = "; ".join(x for x in [reason, extra_reason, "web_identity_strong"] if x)
        return out

    out = dict(evaluation)
    out["identity_score"] = min(out.get("identity_score", 0), 84)
    out["identity_verdict"] = "review" if out["identity_score"] >= 50 else "rejected"
    out["identity_reason"] = "; ".join(
        x for x in [reason, extra_reason, "web_search_identity_gate"] if x
    )
    return out


def _apply_bing_maps_identity_gate(lead: Lead, candidate: dict, evaluation: dict) -> dict:
    """
    Bing Maps is map evidence, but still new in this app. Require accepted
    deterministic identity and at least state-level map-location evidence
    before it can become a final URL.
    """
    if candidate.get("source") != "bing_maps":
        return evaluation
    if candidate.get("diagnostic") or not candidate.get("url"):
        return evaluation
    if _is_provider_infrastructure_url(candidate.get("url")):
        out = dict(evaluation)
        out["identity_score"] = 0
        out["identity_verdict"] = "rejected"
        reason = out.get("identity_reason") or ""
        out["identity_reason"] = "; ".join(
            x for x in [reason, "provider_infrastructure_url", "bing_maps_identity_gate"] if x
        )
        return out

    location_level = evaluation.get("location_match_level")
    company_score = evaluation.get("company_match_score") or 0
    domain_score = evaluation.get("domain_match_score") or 0
    overlap = business_identity_overlap(lead.company, candidate.get("title"))
    provenance = ((candidate.get("raw") or {}).get("website_provenance") or "")
    has_website_provenance = provenance in {
        "explicit_website_label",
        "icon_action_link",
        "visible_subtitle_url",
    }
    has_location_evidence = location_level in {"exact", "nearby", "state_only"}
    strong_identity = (
        evaluation.get("identity_verdict") == "accepted"
        and has_location_evidence
        and location_level != "contradiction"
        and has_website_provenance
        and domain_score >= 85
        and (company_score >= 45 or (overlap is not None and overlap >= 0.5))
    )

    reason = evaluation.get("identity_reason") or ""
    overlap_label = "none" if overlap is None else round(overlap, 2)
    extra_reason = f"bing_maps_gate; provenance={provenance or 'none'}; business_overlap={overlap_label}"

    if strong_identity:
        out = dict(evaluation)
        out["identity_reason"] = "; ".join(x for x in [reason, extra_reason, "bing_maps_identity_strong"] if x)
        return out

    out = dict(evaluation)
    out["identity_score"] = min(out.get("identity_score", 0), 84)
    out["identity_verdict"] = "review" if out["identity_score"] >= 50 else "rejected"
    out["identity_reason"] = "; ".join(
        x for x in [reason, extra_reason, "bing_maps_identity_gate"] if x
    )
    return out


def _is_provider_infrastructure_url(url: str | None) -> bool:
    host = urlparse(url or "").netloc.lower()
    if host.startswith("www."):
        host = host[4:]
    blocked = (
        "bing.com",
        "bingplaces.com",
        "microsoft.com",
        "virtualearth.net",
        "msn.com",
        "maplibre.org",
        "cookieyes.com",
        "facebook.com",
        "instagram.com",
        "linkedin.com",
        "twitter.com",
        "x.com",
        "yelp.com",
        "clarity.ms",
        "doubleclick.net",
        "googletagmanager.com",
    )
    return not host or any(part in host for part in blocked)


def _candidate_is_promotable(evaluation: dict, candidate: dict | None = None) -> bool:
    verdict = evaluation.get("identity_verdict")
    location_level = evaluation.get("location_match_level")
    if (
        candidate
        and candidate.get("source") == "gmaps"
        and candidate.get("mode") != "coordinate_company"
        and verdict == "review"
    ):
        return False
    if (
        candidate
        and candidate.get("source") in {"web_bing", "web_duckduckgo"}
        and verdict == "review"
    ):
        return False
    if candidate and candidate.get("source") == "bing_maps" and verdict == "review":
        return False
    return verdict == "accepted" or (verdict == "review" and location_level != "contradiction")


def _has_promotable_search_candidate(candidates: list[dict]) -> bool:
    return any(_candidate_is_promotable(c.get("evaluation") or {}, c) for c in candidates)


def _search_candidate_sort_key(candidate: dict) -> tuple:
    evaluation = candidate.get("evaluation") or {}
    return (
        1 if _candidate_is_promotable(evaluation, candidate) else 0,
        1 if evaluation.get("identity_verdict") == "accepted" else 0,
        evaluation.get("identity_score", 0),
        1 if candidate.get("source") == "gmaps" else 0,
        1 if candidate.get("source") == "bing_maps" else 0,
        -int(candidate.get("rank") or 999),
    )


def _best_search_candidate(candidates: list[dict], sources: set[str] | None = None) -> dict | None:
    filtered = [
        c for c in candidates
        if c.get("url") and not c.get("diagnostic") and (sources is None or c.get("source") in sources)
    ]
    if not filtered:
        return None
    return max(filtered, key=_search_candidate_sort_key)


def _result_candidate_is_promotable(item: tuple) -> bool:
    source, _, evaluation, meta = item
    if (
        source == "gmaps"
        and meta
        and meta.get("mode") != "coordinate_company"
        and evaluation.get("identity_verdict") == "review"
    ):
        return False
    if source in {"web_bing", "web_duckduckgo"} and evaluation.get("identity_verdict") == "review":
        return False
    if source == "bing_maps" and evaluation.get("identity_verdict") == "review":
        return False
    return evaluation.get("identity_verdict") in {"accepted", "review"}


def _web_search_diagnostics(candidates: list[dict]) -> dict:
    web = [c for c in candidates if c.get("source") in {"web_bing", "web_duckduckgo"}]
    diagnostics = [c for c in web if c.get("diagnostic")]
    errors = [c for c in diagnostics if c.get("error")]
    return {
        "attempted": bool(web),
        "provider_count": len({c.get("source") for c in web if c.get("source")}),
        "query_count": len({(c.get("source"), c.get("query")) for c in web if c.get("query")}),
        "diagnostic_count": len(diagnostics),
        "error_count": len(errors),
        "parsed_count": sum(c.get("parsed_count") or 0 for c in diagnostics),
        "candidate_count": sum(1 for c in web if c.get("url") and not c.get("diagnostic")),
        "error_summary": "; ".join(
            f"{c.get('source')} {c.get('mode')}: {c.get('error')}"
            for c in errors[:3]
        ),
    }


def _candidate_key(candidate: dict) -> tuple:
    return (
        candidate.get("source"),
        candidate.get("mode"),
        candidate.get("query"),
        candidate.get("rank"),
        (candidate.get("url") or "").rstrip("/"),
    )


# ---------------------------------------------------------------------------
# Result builder
# ---------------------------------------------------------------------------
def _build_result(ctx: dict) -> dict:
    result_scoring_start = time.time()
    lead   = ctx["lead"]
    scrape = ctx["scrape"]
    addr   = ctx["addr_parts"]

    haiku_i = ctx.get("haiku_initial") or {}
    perp    = ctx.get("perplexity") or {}
    haiku_f = ctx.get("haiku_final") or {}
    manual  = ctx.get("manual") or {}
    hist    = ctx.get("historical") or {}
    search_candidates = ctx.get("search_candidates") or []
    web_diag = _web_search_diagnostics(search_candidates)
    gmaps_best = _best_search_candidate(search_candidates, {"gmaps"})
    bing_maps_best = _best_search_candidate(search_candidates, {"bing_maps"})
    web_best = _best_search_candidate(search_candidates, {"web_bing", "web_duckduckgo"})
    if gmaps_best and _candidate_is_promotable(gmaps_best.get("evaluation") or {}, gmaps_best):
        scrape = dict(scrape)
        raw = gmaps_best.get("raw") or {}
        scrape.update({
            "website": raw.get("website") or gmaps_best.get("url"),
            "gmaps_listing_name": raw.get("gmaps_listing_name") or gmaps_best.get("title"),
            "address": raw.get("address") or gmaps_best.get("address_or_snippet"),
            "phone": raw.get("phone") or gmaps_best.get("phone"),
            "gmaps_url": raw.get("gmaps_url") or gmaps_best.get("maps_url"),
            "location_match": raw.get("location_match") or gmaps_best.get("location_match"),
            "found": raw.get("found", scrape.get("found")),
        })
        addr = parse_gmaps_address(scrape.get("address") or "")

    g_score  = ctx.get("gmaps_score")
    h_score  = haiku_i.get("score")
    hf_score = haiku_f.get("score")

    # Candidate URLs are scored deterministically. Provider confidence does not
    # drive final confidence; AI output only contributes structured evidence.
    p_url     = perp.get("url")
    maps_url  = scrape.get("website")
    maps_eval = (gmaps_best or {}).get("evaluation") or score_candidate_url(
        source_company       = lead.company,
        source_city          = lead.city,
        source_state         = lead.state,
        source_country       = lead.country,
        candidate_url        = maps_url,
        candidate_source     = "gmaps",
        candidate_name       = scrape.get("gmaps_listing_name"),
        candidate_address    = scrape.get("address"),
        gmaps_location_match = scrape.get("location_match"),
        provider_confidence  = g_score,
    )
    web_url = (web_best or {}).get("url")
    web_eval = (web_best or {}).get("evaluation") or score_candidate_url(
        source_company=lead.company, source_city=lead.city, source_state=lead.state,
        source_country=lead.country, candidate_url=None, candidate_source=None,
    )
    bing_maps_url = (bing_maps_best or {}).get("url")
    bing_maps_eval = (bing_maps_best or {}).get("evaluation") or score_candidate_url(
        source_company=lead.company, source_city=lead.city, source_state=lead.state,
        source_country=lead.country, candidate_url=None, candidate_source=None,
    )
    perplexity_eval = score_candidate_url(
        source_company       = lead.company,
        source_city          = lead.city,
        source_state         = lead.state,
        source_country       = lead.country,
        candidate_url        = p_url,
        candidate_source     = "perplexity",
        candidate_name       = perp.get("official_name"),
        candidate_address    = None,
        evidence_location    = perp.get("evidence_location"),
        evidence_url         = perp.get("evidence_url"),
        evidence_company_match  = perp.get("company_match"),
        evidence_location_match = perp.get("location_match"),
        evidence_is_official    = perp.get("is_official"),
    )
    hist_best = hist.get("best") or {}
    manual_best = manual.get("best") or {}
    manual_candidate = manual_best.get("candidate")
    manual_eval = manual_best.get("evaluation") or score_candidate_url(
        source_company=lead.company, source_city=lead.city, source_state=lead.state,
        source_country=lead.country, candidate_url=None, candidate_source=None,
    )
    manual_url = getattr(manual_candidate, "website", None)
    hist_candidate = hist_best.get("candidate")
    hist_eval = hist_best.get("evaluation") or score_candidate_url(
        source_company=lead.company, source_city=lead.city, source_state=lead.state,
        source_country=lead.country, candidate_url=None, candidate_source=None,
    )
    legacy_diag = (hist.get("diagnostics") or {}).get("legacy_db", {})
    hist_url = getattr(hist_candidate, "website", None)
    candidates = []
    if maps_url:
        candidates.append(("gmaps", maps_url, maps_eval, gmaps_best))
    if bing_maps_url:
        candidates.append(("bing_maps", bing_maps_url, bing_maps_eval, bing_maps_best))
    if web_url:
        candidates.append((web_best.get("source", "web_search"), web_url, web_eval, web_best))
    if p_url:
        candidates.append(("perplexity", p_url, perplexity_eval, None))
    if manual_url:
        candidates.append(("manual_verified", manual_url, manual_eval, None))
    if hist_url:
        candidates.append((getattr(hist_candidate, "source", "historical"), hist_url, hist_eval, None))

    best_source, best_url, best_eval, best_meta = (None, None, score_candidate_url(
        source_company=lead.company, source_city=lead.city, source_state=lead.state,
        source_country=lead.country, candidate_url=None, candidate_source=None,
    ), None)
    if candidates:
        best_source, best_url, best_eval, best_meta = max(
            candidates,
            key=lambda item: (
                1 if _result_candidate_is_promotable(item) else 0,
                item[2].get("identity_score", 0),
                1 if (
                    item[0] == "perplexity"
                    and item[2].get("identity_verdict") in {"accepted", "review"}
                    and (perp.get("company_match") or "").lower() in {"exact", "dba", "acronym"}
                    and (perp.get("is_official") is True)
                ) else 0,
                1 if item[0] == "manual_verified" else 0,
                1 if item[0] == "gmaps" else 0,
            ),
        )

    candidate_score = best_eval.get("identity_score", 0)
    candidate_verdict = best_eval.get("identity_verdict")
    candidate_location_level = best_eval.get("location_match_level")
    historical_best = best_source in {"salesforce", "legacy_db"}
    retry_gmaps_best = (
        best_source == "gmaps"
        and (best_meta or {}).get("mode") not in (None, "", "coordinate_company")
    )
    web_search_best = best_source in {"web_bing", "web_duckduckgo"}
    bing_maps_search_best = best_source == "bing_maps"
    review_is_final = (
        candidate_verdict == "review"
        and candidate_location_level != "contradiction"
        and not historical_best
        and not retry_gmaps_best
        and not web_search_best
        and not bing_maps_search_best
    )
    final_url = best_url if candidate_verdict == "accepted" or review_is_final else None
    if final_url:
        final_score = candidate_score
        final_verdict = candidate_verdict
        final_reason = best_eval.get("identity_reason")
        final_company_score = best_eval.get("company_match_score")
        final_domain_score = best_eval.get("domain_match_score")
        final_location_level = candidate_location_level
    else:
        final_score = 0
        final_verdict = "no_candidate" if not candidates else "rejected"
        final_reason = (
            "No candidate URL cleared final promotion threshold."
            if candidates else "No candidate URL was returned."
        )
        final_company_score = 0
        final_domain_score = 0
        final_location_level = "unknown"

    if final_verdict == "accepted" and final_url:
        tier = "High"
    elif final_verdict == "review" and final_url:
        tier = "Medium"
    else:
        tier = None
    url_source = best_source if final_url else None
    url_changed = bool(final_url and p_url and maps_url and final_url.rstrip("/") != maps_url.rstrip("/"))
    selected_search_candidate = None
    if final_url and url_source in {"gmaps", "bing_maps", "web_bing", "web_duckduckgo"}:
        selected_key = _candidate_key(best_meta) if best_meta else None
        if selected_key:
            for candidate in search_candidates:
                candidate["selected"] = _candidate_key(candidate) == selected_key
                if candidate["selected"]:
                    selected_search_candidate = candidate
        else:
            for candidate in search_candidates:
                candidate["selected"] = False
                if candidate.get("source") == url_source and (candidate.get("url") or "").rstrip("/") == final_url.rstrip("/"):
                    selected_search_candidate = candidate
            if selected_search_candidate:
                selected_key = _candidate_key(selected_search_candidate)
                for candidate in search_candidates:
                    candidate["selected"] = _candidate_key(candidate) == selected_key
    else:
        for candidate in search_candidates:
            candidate["selected"] = False

    # Status
    found = scrape.get("found", False)
    if not candidates:
        status = "no_website" if found else "not_found"
    elif not final_url:
        status = "rejected_candidate"
    elif url_source == "manual_verified":
        status = "manual_verified"
    elif url_source in {"salesforce", "legacy_db"}:
        status = "historical_found"
    elif url_source == "perplexity" and not maps_url:
        status = "perplexity_found"
    elif url_source == "bing_maps":
        status = "bing_maps_found"
    elif url_source in {"web_bing", "web_duckduckgo"}:
        status = "web_search_found"
    elif tier == "High":
        status = "confirmed"
    elif tier == "Medium":
        status = "medium_confidence"
    elif tier == "Low":
        status = "low_confidence"
    else:
        status = "low_confidence"

    elapsed_ms = int((time.time() - ctx.get("start_time", time.time())) * 1000)
    timings = ctx.get("timings") or {}
    total_cost = (ctx.get("haiku_initial_cost", 0) +
                  ctx.get("perplexity_cost", 0) +
                  ctx.get("haiku_final_cost", 0))

    sigs = haiku_i.get("signals") or haiku_f.get("signals") or {}
    timings["scoring_latency_ms"] = timings.get("scoring_latency_ms", 0) + int((time.time() - result_scoring_start) * 1000)

    return {
        # Input
        "company":              lead.company,
        "city":                 lead.city,
        "state":                lead.state,
        "country":              lead.country,
        "source_sheet":         lead.source_sheet,
        "source_excel_row":     getattr(lead, "original_row_idx", None),
        "source_valid_index":   getattr(lead, "source_valid_index", None),
        # GMaps
        "gmaps_found":          found,
        "gmaps_listing_name":   scrape.get("gmaps_listing_name"),
        "gmaps_website":        maps_url,
        "gmaps_phone":          scrape.get("phone"),
        "gmaps_address":        scrape.get("address"),
        "gmaps_street":         addr.get("gmaps_street"),
        "gmaps_city_scraped":   addr.get("gmaps_city_scraped"),
        "gmaps_state_scraped":  addr.get("gmaps_state_scraped"),
        "gmaps_zip":            addr.get("gmaps_zip"),
        "gmaps_location_match": scrape.get("location_match"),
        # Scores
        "name_similarity":         round(name_similarity(lead.company, scrape.get("gmaps_listing_name")) * 100),
        "gmaps_confidence_score":  g_score,
        # Haiku initial signals
        "sig_site_name":        sigs.get("name_match"),
        "sig_site_location":    sigs.get("location_found"),
        "sig_isn_mention":      sigs.get("isn_mention"),
        "sig_disqualifier":     sigs.get("disqualifier"),
        "haiku_initial_confidence":  h_score,
        "haiku_initial_match":       haiku_i.get("match"),
        "haiku_initial_stop_reason": haiku_i.get("stop_reason"),
        "haiku_initial_latency_ms":  haiku_i.get("latency_ms"),
        # Perplexity
        "perplexity_url":       p_url,
        "perplexity_reason":    perp.get("reason"),
        "perplexity_official_name":     perp.get("official_name"),
        "perplexity_evidence_location": perp.get("evidence_location"),
        "perplexity_evidence_url":      perp.get("evidence_url"),
        "perplexity_company_match":     perp.get("company_match"),
        "perplexity_location_match":    perp.get("location_match"),
        "perplexity_is_official":       perp.get("is_official"),
        "perplexity_reject_reason":     perp.get("reject_reason"),
        "perplexity_citations": json.dumps(perp.get("citations", [])),
        "perplexity_latency_ms":perp.get("latency_ms"),
        # Manual verified associations
        "manual_url":        manual_url,
        "manual_association_id": getattr(manual_candidate, "association_id", None) if manual_candidate else None,
        "manual_association_type": getattr(manual_candidate, "association_type", None) if manual_candidate else None,
        "manual_notes":      getattr(manual_candidate, "notes", None) if manual_candidate else None,
        "manual_verified_by": getattr(manual_candidate, "verified_by", None) if manual_candidate else None,
        "manual_verified_at": getattr(manual_candidate, "verified_at", None) if manual_candidate else None,
        "manual_identity_score": manual_eval.get("identity_score"),
        "manual_identity_verdict": manual_eval.get("identity_verdict"),
        "manual_identity_reason": manual_eval.get("identity_reason"),
        # Historical enrichment
        "historical_url":        hist_url,
        "historical_source":     getattr(hist_candidate, "source", None) if hist_candidate else None,
        "historical_record_type":getattr(hist_candidate, "record_type", None) if hist_candidate else None,
        "historical_record_id":  getattr(hist_candidate, "record_id", None) if hist_candidate else None,
        "historical_name":       getattr(hist_candidate, "name", None) if hist_candidate else None,
        "historical_city":       getattr(hist_candidate, "city", None) if hist_candidate else None,
        "historical_state":      getattr(hist_candidate, "state", None) if hist_candidate else None,
        "historical_raw_source": getattr(hist_candidate, "raw_source", None) if hist_candidate else None,
        "historical_candidate_count": hist.get("candidates_found", 0),
        "historical_legacy_raw_rows": legacy_diag.get("raw_rows"),
        "historical_legacy_rows_with_email": legacy_diag.get("rows_with_email"),
        "historical_legacy_usable_domains": legacy_diag.get("usable_domains"),
        "historical_legacy_filtered_domains": legacy_diag.get("filtered_domains"),
        "historical_legacy_query_name_like": legacy_diag.get("query_name_like"),
        "historical_legacy_query_state": legacy_diag.get("query_state"),
        "historical_legacy_query_state_abbr": legacy_diag.get("query_state_abbr"),
        "historical_legacy_query_state_full": legacy_diag.get("query_state_full"),
        "historical_identity_score":  hist_eval.get("identity_score"),
        "historical_identity_verdict": hist_eval.get("identity_verdict"),
        "historical_identity_reason": hist_eval.get("identity_reason"),
        "historical_errors":     "; ".join(hist.get("errors") or []),
        "historical_latency_ms": hist.get("latency_ms"),
        # Haiku validation
        "haiku_final_confidence":    hf_score,
        "haiku_final_match":         haiku_f.get("match"),
        "haiku_final_stop_reason":   haiku_f.get("stop_reason"),
        "haiku_final_latency_ms":    haiku_f.get("latency_ms"),
        # Final
        "final_url":            final_url,
        "url_source":           url_source,
        "url_changed":          url_changed,
        "final_confidence_score":final_score,
        "confidence_tier":      tier,
        "identity_verdict":     final_verdict,
        "identity_reason":      final_reason,
        "company_match_score":  final_company_score,
        "domain_match_score":   final_domain_score,
        "location_match_level": final_location_level,
        "gmaps_identity_score": maps_eval.get("identity_score"),
        "gmaps_identity_verdict": maps_eval.get("identity_verdict"),
        "perplexity_identity_score": perplexity_eval.get("identity_score"),
        "perplexity_identity_verdict": perplexity_eval.get("identity_verdict"),
        "manual_best_identity_score": manual_eval.get("identity_score"),
        "manual_best_identity_verdict": manual_eval.get("identity_verdict"),
        "historical_best_identity_score": hist_eval.get("identity_score"),
        "historical_best_identity_verdict": hist_eval.get("identity_verdict"),
        "haiku_reasoning":      (haiku_f.get("signals") or haiku_i.get("signals") or {}).get("reasoning", ""),
        "stages_run":           ",".join(ctx.get("stages_run", ["gmaps"])),
        "total_cost_usd":       round(total_cost, 6),
        "total_latency_ms":     elapsed_ms,
        "gmaps_latency_ms":     timings.get("gmaps_latency_ms"),
        "bing_maps_latency_ms": timings.get("bing_maps_latency_ms"),
        "web_search_latency_ms": timings.get("web_search_latency_ms"),
        "manual_latency_ms":    timings.get("manual_latency_ms"),
        "scoring_latency_ms":   timings.get("scoring_latency_ms"),
        "scrape_attempts":      scrape.get("gmaps_attempts") or 1,
        "search_candidates_evaluated": len(search_candidates),
        "gmaps_attempts":       scrape.get("gmaps_attempts") or 1,
        "web_search_attempted": web_diag["attempted"],
        "web_search_provider_count": web_diag["provider_count"],
        "web_search_query_count": web_diag["query_count"],
        "web_search_candidate_count": web_diag["candidate_count"],
        "web_search_diagnostic_count": web_diag["diagnostic_count"],
        "web_search_error_count": web_diag["error_count"],
        "web_search_parsed_count": web_diag["parsed_count"],
        "web_search_error_summary": web_diag["error_summary"],
        "selected_candidate_source": (selected_search_candidate or {}).get("source"),
        "selected_candidate_mode":   (selected_search_candidate or {}).get("mode"),
        "selected_candidate_query":  (selected_search_candidate or {}).get("query"),
        "selected_candidate_rank":   (selected_search_candidate or {}).get("rank"),
        "search_candidates":    search_candidates,
        "status":               status,
        "error":                scrape.get("error"),
    }


def _result_to_db_row(job_id: str, r: dict) -> dict:
    keys = [
        "company","city","state","country","source_sheet","source_excel_row","source_valid_index",
        "gmaps_found","gmaps_listing_name","gmaps_website","gmaps_phone",
        "gmaps_address","gmaps_street","gmaps_city_scraped","gmaps_state_scraped","gmaps_zip",
        "gmaps_location_match","name_similarity","gmaps_confidence_score",
        "sig_site_name","sig_site_location","sig_isn_mention","sig_disqualifier",
        "haiku_initial_confidence","haiku_initial_match","haiku_initial_stop_reason","haiku_initial_latency_ms",
        "perplexity_url","perplexity_reason",
        "perplexity_official_name","perplexity_evidence_location","perplexity_evidence_url",
        "perplexity_company_match","perplexity_location_match","perplexity_is_official",
        "perplexity_reject_reason","perplexity_citations","perplexity_latency_ms",
        "manual_url","manual_association_id","manual_association_type","manual_notes",
        "manual_verified_by","manual_verified_at","manual_identity_score",
        "manual_identity_verdict","manual_identity_reason",
        "historical_url","historical_source","historical_record_type","historical_record_id",
        "historical_name","historical_city","historical_state","historical_raw_source",
        "historical_candidate_count","historical_legacy_raw_rows",
        "historical_legacy_rows_with_email","historical_legacy_usable_domains",
        "historical_legacy_filtered_domains","historical_legacy_query_name_like",
        "historical_legacy_query_state","historical_legacy_query_state_abbr",
        "historical_legacy_query_state_full","historical_identity_score","historical_identity_verdict",
        "historical_identity_reason","historical_errors","historical_latency_ms",
        "haiku_final_confidence","haiku_final_match","haiku_final_stop_reason","haiku_final_latency_ms",
        "final_url","url_source","url_changed","final_confidence_score","confidence_tier",
        "identity_verdict","identity_reason","company_match_score","domain_match_score",
        "location_match_level","gmaps_identity_score","gmaps_identity_verdict",
        "perplexity_identity_score","perplexity_identity_verdict",
        "manual_best_identity_score","manual_best_identity_verdict",
        "historical_best_identity_score","historical_best_identity_verdict",
        "haiku_reasoning","stages_run","total_cost_usd","total_latency_ms",
        "gmaps_latency_ms","bing_maps_latency_ms","web_search_latency_ms",
        "manual_latency_ms","scoring_latency_ms","scrape_attempts",
        "search_candidates_evaluated","gmaps_attempts","web_search_candidate_count",
        "web_search_attempted","web_search_provider_count","web_search_query_count",
        "web_search_diagnostic_count","web_search_error_count",
        "web_search_parsed_count","web_search_error_summary",
        "selected_candidate_source","selected_candidate_mode","selected_candidate_query","selected_candidate_rank",
        "status",
    ]
    row = {"job_id": job_id}
    for k in keys:
        row[k] = r.get(k)
    for k in (
        "gmaps_found",
        "haiku_initial_match",
        "perplexity_is_official",
        "haiku_final_match",
        "url_changed",
        "web_search_attempted",
    ):
        row[k] = _coerce_db_bool(row.get(k))
    return row


def _coerce_db_bool(value):
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        text = value.strip().lower()
        if text in {"true", "1", "yes", "y"}:
            return True
        if text in {"false", "0", "no", "n", ""}:
            return False
    return None


def _search_candidate_to_db_row(job_id: str, result_id: int, result: dict,
                                candidate: dict) -> dict:
    evaluation = candidate.get("evaluation") or {}
    return {
        "job_id": job_id,
        "result_id": result_id,
        "company": result.get("company"),
        "city": result.get("city"),
        "state": result.get("state"),
        "country": result.get("country"),
        "source": candidate.get("source"),
        "mode": candidate.get("mode"),
        "query": candidate.get("query"),
        "rank": candidate.get("rank"),
        "title": candidate.get("title"),
        "url": candidate.get("url"),
        "address_or_snippet": candidate.get("address_or_snippet"),
        "phone": candidate.get("phone"),
        "maps_url": candidate.get("maps_url"),
        "identity_score": evaluation.get("identity_score"),
        "identity_verdict": evaluation.get("identity_verdict"),
        "identity_reason": evaluation.get("identity_reason"),
        "company_match_score": evaluation.get("company_match_score"),
        "domain_match_score": evaluation.get("domain_match_score"),
        "location_match_level": evaluation.get("location_match_level"),
        "selected": bool(candidate.get("selected")),
        "error": candidate.get("error"),
        "diagnostic": bool(candidate.get("diagnostic")),
        "http_status": candidate.get("http_status"),
        "response_bytes": candidate.get("response_bytes"),
        "parsed_count": candidate.get("parsed_count"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }


def _cap_text(value, limit: int):
    if value is None:
        return None
    text_value = str(value)
    if len(text_value) <= limit:
        return text_value
    return text_value[:limit] + "...[truncated]"


def _cap_search_candidate_db_row(row: dict) -> dict:
    """Keep candidate audit rows bounded so persistence cannot strand a job."""
    caps = {
        "query": 500,
        "title": 500,
        "url": 1200,
        "address_or_snippet": 1200,
        "phone": 100,
        "maps_url": 1200,
        "identity_reason": 1200,
        "error": 800,
    }
    out = dict(row)
    for key, limit in caps.items():
        out[key] = _cap_text(out.get(key), limit)
    return out


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
COLUMNS = [
    "source_valid_index","source_sheet","source_excel_row",
    "company","city","state","country",
    "gmaps_found","gmaps_listing_name","gmaps_website","gmaps_phone",
    "gmaps_address","gmaps_street","gmaps_city_scraped","gmaps_state_scraped","gmaps_zip",
    "gmaps_location_match",
    "name_similarity","gmaps_confidence_score","gmaps_identity_score","gmaps_identity_verdict",
    "sig_site_name","sig_site_location","sig_isn_mention","sig_disqualifier",
    "haiku_initial_confidence","haiku_initial_match","haiku_initial_stop_reason",
    "perplexity_url","perplexity_identity_score","perplexity_identity_verdict",
    "perplexity_reason","perplexity_official_name","perplexity_evidence_location","perplexity_evidence_url",
    "perplexity_company_match","perplexity_location_match","perplexity_is_official",
    "perplexity_reject_reason","perplexity_citations",
    "manual_url","manual_association_id","manual_association_type","manual_notes",
    "manual_verified_by","manual_verified_at","manual_identity_score",
    "manual_identity_verdict","manual_identity_reason",
    "historical_url","historical_source","historical_record_type","historical_record_id",
    "historical_name","historical_city","historical_state","historical_raw_source",
    "historical_candidate_count","historical_legacy_raw_rows",
    "historical_legacy_rows_with_email","historical_legacy_usable_domains",
    "historical_legacy_filtered_domains","historical_legacy_query_name_like",
    "historical_legacy_query_state","historical_legacy_query_state_abbr",
    "historical_legacy_query_state_full","historical_identity_score","historical_identity_verdict",
    "historical_identity_reason","historical_errors","historical_latency_ms",
    "haiku_final_confidence","haiku_final_match","haiku_final_stop_reason",
    "final_url","url_source","url_changed","final_confidence_score","confidence_tier",
    "identity_verdict","identity_reason","company_match_score","domain_match_score","location_match_level",
    "haiku_reasoning","stages_run","total_cost_usd","total_latency_ms",
    "gmaps_latency_ms","bing_maps_latency_ms","web_search_latency_ms",
    "manual_latency_ms","scoring_latency_ms","scrape_attempts",
    "search_candidates_evaluated","gmaps_attempts","web_search_candidate_count",
    "web_search_attempted","web_search_provider_count","web_search_query_count",
    "web_search_diagnostic_count","web_search_error_count",
    "web_search_parsed_count","web_search_error_summary",
    "selected_candidate_source","selected_candidate_mode","selected_candidate_query","selected_candidate_rank",
    "status",
]


def _write_output(job: Job) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = Path(__file__).parent.parent / "data" / "outputs" / f"results_{job.job_id}_{ts}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # ── Results sheet ────────────────────────────────────────────────────
    ws = wb.active; ws.title = "Results"
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    for ci, col in enumerate(COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal="center")

    tier_fills = {
        "High":   PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
        "Medium": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Low":    PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    }
    for ri, result in enumerate(job.results, 2):
        fill = tier_fills.get(result.get("confidence_tier", ""))
        for ci, col in enumerate(COLUMNS, 1):
            val = result.get(col, "")
            if isinstance(val, bool): val = "Yes" if val else "No"
            c = ws.cell(row=ri, column=ci, value=val)
            if fill: c.fill = fill

    widths = {"company":40,"gmaps_website":36,"final_url":36,"gmaps_address":40,
               "identity_reason":55,"haiku_reasoning":55,"perplexity_reason":55,
               "perplexity_citations":50,"gmaps_street":30}
    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, max(len(col)+2, 12))
    ws.freeze_panes = "A2"
    _write_candidates_sheet(wb, job)
    _write_stage_events_sheet(wb, job)

    # ── Report sheet ─────────────────────────────────────────────────────
    if job.report:
        _write_report_sheet(wb, job.report)

    wb.save(str(out))
    return str(out)


def _write_candidates_sheet(wb, job: Job):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cols = [
        "company", "city", "state", "country", "selected", "diagnostic",
        "source", "mode", "query", "rank",
        "title", "url", "address_or_snippet", "phone", "maps_url",
        "http_status", "response_bytes", "parsed_count",
        "identity_score", "identity_verdict", "identity_reason",
        "company_match_score", "domain_match_score", "location_match_level",
        "error",
    ]
    ws = wb.create_sheet("Candidates")
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    selected_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    row_idx = 2
    for result in job.results:
        for candidate in result.get("search_candidates", []):
            evaluation = candidate.get("evaluation") or {}
            row = {
                "company": result.get("company"),
                "city": result.get("city"),
                "state": result.get("state"),
                "country": result.get("country"),
                "selected": "Yes" if candidate.get("selected") else "No",
                "diagnostic": "Yes" if candidate.get("diagnostic") else "No",
                "source": candidate.get("source"),
                "mode": candidate.get("mode"),
                "query": candidate.get("query"),
                "rank": candidate.get("rank"),
                "title": candidate.get("title"),
                "url": candidate.get("url"),
                "address_or_snippet": candidate.get("address_or_snippet"),
                "phone": candidate.get("phone"),
                "maps_url": candidate.get("maps_url"),
                "http_status": candidate.get("http_status"),
                "response_bytes": candidate.get("response_bytes"),
                "parsed_count": candidate.get("parsed_count"),
                "identity_score": evaluation.get("identity_score"),
                "identity_verdict": evaluation.get("identity_verdict"),
                "identity_reason": evaluation.get("identity_reason"),
                "company_match_score": evaluation.get("company_match_score"),
                "domain_match_score": evaluation.get("domain_match_score"),
                "location_match_level": evaluation.get("location_match_level"),
                "error": candidate.get("error"),
            }
            for ci, col in enumerate(cols, 1):
                c = ws.cell(row=row_idx, column=ci, value=row.get(col, ""))
                if candidate.get("selected"):
                    c.fill = selected_fill
            row_idx += 1

    widths = {
        "company": 38,
        "query": 42,
        "title": 36,
        "url": 42,
        "address_or_snippet": 55,
        "maps_url": 48,
        "error": 42,
        "identity_reason": 48,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, max(len(col) + 2, 12))
    ws.freeze_panes = "A2"


def _write_stage_events_sheet(wb, job: Job):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    cols = [
        "created_at", "event_type", "stage", "company", "city", "state",
        "source_sheet", "worker", "elapsed_ms", "is_active_snapshot",
        "message", "details_json",
    ]
    ws = wb.create_sheet("Worker Events")
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")

    for ci, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")

    for ri, event in enumerate(job.stage_events_snapshot(limit=0), 2):
        for ci, col in enumerate(cols, 1):
            ws.cell(row=ri, column=ci, value=event.get(col))

    widths = {
        "created_at": 28,
        "event_type": 24,
        "stage": 24,
        "company": 42,
        "worker": 20,
        "message": 48,
        "details_json": 70,
    }
    for ci, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, max(len(col) + 2, 12))
    ws.freeze_panes = "A2"


def _write_report_sheet(wb, report: dict):
    from openpyxl.styles import Font, PatternFill
    rs = wb.create_sheet("Job Report")
    hf = Font(bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")

    def section(title, rows):
        rs.append([])
        c = rs.cell(row=rs.max_row+1, column=1, value=title)
        c.font = hf; c.fill = hfill
        for label, value in rows:
            rs.append([label, value])

    section("JOB", [
        ("Job ID",           report["job_id"]),
        ("Experiment",       report.get("experiment_name","")),
        ("Run Date",         report["run_date"]),
        ("Duration (sec)",   report["duration_sec"]),
        ("File",             report["input"]["file_name"]),
        ("Valid Rows",       report["input"]["valid_rows"]),
    ])
    source = report.get("source", {})
    section("SOURCE RANGE", [
        ("File hash",             source.get("file_hash")),
        ("Raw rows in source",    source.get("total_rows_raw")),
        ("Valid rows in source",  source.get("total_valid_rows")),
        ("Valid row start",       source.get("valid_row_start")),
        ("Valid row end",         source.get("valid_row_end")),
        ("Processed valid rows",  source.get("processed_valid_rows")),
    ])
    pip = report.get("pipeline", {})
    section("PIPELINE CONFIG", [
        ("Haiku enabled",            pip.get("haiku_enabled")),
        ("Haiku validation enabled", pip.get("haiku_validation_enabled")),
        ("Perplexity enabled",       pip.get("perplexity_enabled")),
        ("Perplexity trigger",       pip.get("perplexity_trigger")),
        ("Manual associations",      pip.get("manual_associations_enabled")),
        ("Salesforce enrichment",    pip.get("salesforce_enrichment_enabled")),
        ("Legacy DB enrichment",     pip.get("legacy_enrichment_enabled")),
        ("Web search fallback",      pip.get("web_search_fallback_enabled")),
        ("Bing web fallback",        pip.get("bing_web_enabled")),
        ("DuckDuckGo web fallback",  pip.get("duckduckgo_web_enabled")),
        ("Bing Maps fallback",       pip.get("bing_maps_enabled")),
    ])
    sett = report.get("job_settings", {})
    section("JOB SETTINGS", [
        ("Scrape workers",              sett.get("scrape_workers")),
        ("Validate workers",            sett.get("validate_workers")),
        ("Stale timeout (sec)",         sett.get("job_stall_timeout_seconds")),
        ("Jina timeout (sec)",          sett.get("jina_timeout_seconds")),
        ("Max scrape errors",           sett.get("max_scrape_errors")),
        ("Maps candidates per mode",    sett.get("gmaps_max_candidates_per_mode")),
        ("Maps early-stop score",       sett.get("gmaps_strong_stop_score")),
        ("Web search max results",      sett.get("web_search_max_results")),
        ("Historical min score",        sett.get("historical_enrichment_min_score")),
    ])
    runtime = report.get("runtime", {})
    section("RUNTIME DIAGNOSTICS", [
        ("App started at",              runtime.get("app_started_at")),
        ("Run started at",             runtime.get("run_started_at")),
        ("Run completed at",           runtime.get("run_completed_at")),
        ("Run duration (sec)",         runtime.get("run_duration_sec")),
        ("Last progress at",           runtime.get("last_progress_at")),
        ("Last heartbeat at",          runtime.get("last_heartbeat_at")),
        ("Max heartbeat gap (sec)",    runtime.get("max_heartbeat_gap_sec")),
        ("Suspected sleep events",     json.dumps(runtime.get("suspected_sleep_events") or [])),
        ("Stale detected",             runtime.get("stale_detected")),
        ("Stale reason",               runtime.get("stale_reason")),
        ("Peak browser count",         runtime.get("peak_browser_count")),
        ("Worker errors",              runtime.get("worker_error_count")),
        ("Active rows at close",        runtime.get("active_rows_at_close_count")),
        ("Process ID",                 runtime.get("process_id")),
        ("Host name",                  runtime.get("host_name")),
        ("Platform",                   runtime.get("platform")),
        ("Python version",             runtime.get("python_version")),
        ("Working directory",          runtime.get("cwd")),
    ])
    gm = report["google_maps"]
    section("GOOGLE MAPS", [
        ("Found listing",    gm["found_listing"]),
        ("Not found",        gm["not_found"]),
        ("Has website",      gm["has_website"]),
        ("No website on listing", gm["no_website_on_listing"]),
        ("Location exact",   gm["location_match_exact"]),
        ("Location partial", gm["location_match_partial"]),
    ])
    conf = report["confidence"]
    section("CONFIDENCE", [
        ("High",       f"{conf['high']['count']} ({conf['high']['pct']}%)"),
        ("Medium",     f"{conf['medium']['count']} ({conf['medium']['pct']}%)"),
        ("Low",        f"{conf['low']['count']} ({conf['low']['pct']}%)"),
        ("Unresolved", f"{conf['unresolved']['count']} ({conf['unresolved']['pct']}%)"),
        ("Avg Score",  conf["avg_score"]),
    ])
    us = report.get("url_sources", {})
    section("URL SOURCES", [
        ("From GMaps",       us.get("from_gmaps", 0)),
        ("From Perplexity",  us.get("from_perplexity", 0)),
        ("From Manual",      us.get("from_manual", 0)),
        ("From Salesforce",  us.get("from_salesforce", 0)),
        ("From Legacy DB",   us.get("from_legacy", 0)),
        ("From Bing Maps",   us.get("from_bing_maps", 0)),
        ("From Bing web",    us.get("from_web_bing", 0)),
        ("From DuckDuckGo web", us.get("from_web_duckduckgo", 0)),
        ("URL Changed",      us.get("url_changed", 0)),
    ])
    sc = report.get("search_candidates", {})
    section("SEARCH CANDIDATES", [
        ("Candidates evaluated", sc.get("evaluated", 0)),
        ("GMaps attempts",       sc.get("gmaps_attempts", 0)),
        ("Recovered by GMaps retry mode", sc.get("gmaps_retry_recovered", 0)),
        ("Recovered by Bing Maps",        sc.get("bing_maps_recovered", 0)),
        ("Recovered by web search",       sc.get("web_search_recovered", 0)),
        ("Web candidates",       sc.get("web_candidates", 0)),
        ("Web attempt rows",     sc.get("web_attempt_rows", 0)),
        ("Web queries",          sc.get("web_queries", 0)),
        ("Web diagnostics",      sc.get("web_diagnostics", 0)),
        ("Web parsed results",   sc.get("web_parsed_results", 0)),
        ("Web errors",           sc.get("web_errors", 0)),
        ("Rejected candidates",  sc.get("rejected", 0)),
    ])
    manual = report.get("manual_associations", {})
    section("MANUAL ASSOCIATIONS", [
        ("Enabled",              manual.get("enabled", True)),
        ("Rows with candidates", manual.get("rows_with_candidates", 0)),
        ("Accepted candidates",  manual.get("accepted_candidates", 0)),
        ("Final URLs",           manual.get("final_urls", 0)),
    ])
    hist = report.get("historical_enrichment", {})
    section("HISTORICAL ENRICHMENT", [
        ("Salesforce enabled",     hist.get("salesforce_enabled", False)),
        ("Legacy DB enabled",      hist.get("legacy_enabled", False)),
        ("Rows with candidates",   hist.get("rows_with_candidates", 0)),
        ("Accepted candidates",    hist.get("accepted_candidates", 0)),
        ("Final URLs from Salesforce", hist.get("from_salesforce", 0)),
        ("Final URLs from Legacy DB",  hist.get("from_legacy", 0)),
        ("Legacy raw SQL rows",        hist.get("legacy_raw_rows", 0)),
        ("Legacy rows with email",     hist.get("legacy_rows_with_email", 0)),
        ("Legacy usable domains",      hist.get("legacy_usable_domains", 0)),
        ("Legacy filtered domains",    hist.get("legacy_filtered_domains", 0)),
        ("Rows with source errors",    hist.get("errors", 0)),
    ])
    ident = report.get("identity", {})
    section("IDENTITY GATE", [
        ("Accepted",         ident.get("accepted", 0)),
        ("Review",           ident.get("review", 0)),
        ("Rejected",         ident.get("rejected", 0)),
        ("No candidate",     ident.get("no_candidate", 0)),
    ])
    h = report["haiku"]
    section("HAIKU", [
        ("Initial calls",    h["calls_initial"]),
        ("Validation calls", h["calls_validation"]),
        ("Input tokens",     h["input_tokens"]),
        ("Output tokens",    h["output_tokens"]),
        ("Truncated responses", h["truncated_responses"]),
        ("Cost",             f"${h['cost_usd']:.6f}"),
    ])
    p = report["perplexity"]
    section("PERPLEXITY", [
        ("Calls",            p["calls"]),
        ("Input tokens",     p["input_tokens"]),
        ("Output tokens",    p["output_tokens"]),
        ("Total citations",  p["total_citations"]),
        ("Cost",             f"${p['cost_usd']:.6f}"),
    ])
    c = report["cost"]
    section("COST", [
        ("Haiku",            f"${c['haiku_usd']:.6f}"),
        ("Perplexity",       f"${c['perplexity_usd']:.6f}"),
        ("Total",            f"${c['total_usd']:.6f}"),
        ("Per record",       f"${c['cost_per_record']:.6f}"),
    ])
    perf = report.get("performance", {})
    section("PERFORMANCE", [
        ("Avg total latency (ms)",      perf.get("avg_total_latency_ms", 0)),
        ("Avg GMaps latency (ms)",      perf.get("avg_gmaps_latency_ms", 0)),
        ("Avg Bing Maps latency (ms)",  perf.get("avg_bing_maps_latency_ms", 0)),
        ("Avg web search latency (ms)", perf.get("avg_web_search_latency_ms", 0)),
        ("Avg manual latency (ms)",     perf.get("avg_manual_latency_ms", 0)),
        ("Avg historical latency (ms)", perf.get("avg_historical_latency_ms", 0)),
        ("Avg scoring latency (ms)",    perf.get("avg_scoring_latency_ms", 0)),
        ("Peak browser count",          perf.get("peak_browser_count", 0)),
    ])

    rs.column_dimensions["A"].width = 28
    rs.column_dimensions["B"].width = 22
